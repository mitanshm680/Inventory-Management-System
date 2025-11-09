"""
Inventory Management System - Backend API
Complete FastAPI backend with all endpoints
"""
from fastapi import FastAPI, Depends, HTTPException, status, Request, UploadFile, File
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from pydantic import BaseModel
from typing import Optional, List, Dict, Any
import json
from datetime import datetime, timedelta
from jose import JWTError, jwt
import hashlib
import logging
import sqlite3
import sys
import os
import csv
import io

# Import our services
from services.inventory_service import InventoryService
from services.user_service import UserService
from utils.logging_config import setup_logging
from database.setup import initialize_database
from database.db_connection import DBConnection

# Setup logging
setup_logging()
logging.info("Starting Inventory Management API")

# Initialize database
initialize_database()

# Access the database connection
db_connection = DBConnection()

# Initialize default admin user with SHA-256 hashed password
try:
    with db_connection.get_cursor() as cursor:
        cursor.execute("SELECT 1")
        # Check if admin exists
        cursor.execute("SELECT password FROM users WHERE username = 'admin'")
        admin = cursor.fetchone()

        if admin:
            # Update existing admin password to SHA-256 if it's plaintext
            if admin['password'] == '1234':
                hashed_password = hashlib.sha256('1234'.encode()).hexdigest()
                cursor.execute("UPDATE users SET password = ? WHERE username = 'admin'", (hashed_password,))
                logging.info("Updated admin password to SHA-256 hash")
        else:
            # Create new admin with hashed password
            hashed_password = hashlib.sha256('1234'.encode()).hexdigest()
            cursor.execute("""
                INSERT INTO users (username, password, role)
                VALUES ('admin', ?, 'admin')
            """, (hashed_password,))
            logging.info("Created admin user with SHA-256 hashed password")
except Exception as e:
    logging.error(f"Database initialization error: {e}")
    sys.exit(1)

# Security settings
SECRET_KEY = "your-secret-key-change-in-production"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 1440  # 24 hours

# FastAPI app
app = FastAPI(
    title="Inventory Management System API",
    description="Complete inventory management with user authentication",
    version="1.0.0"
)

# CORS - Allow both localhost and 127.0.0.1
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:3000",
        "http://127.0.0.1:3000",
        "http://localhost:8001",
        "http://127.0.0.1:8001",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"]
)

# Initialize services
inventory_service = InventoryService()
user_service = UserService()

# ============================================================================
# Pydantic Models
# ============================================================================

class Token(BaseModel):
    access_token: str
    token_type: str

class TokenData(BaseModel):
    username: Optional[str] = None
    role: Optional[str] = None

class User(BaseModel):
    username: str
    role: str

class UserCreate(BaseModel):
    username: str
    password: str
    role: Optional[str] = "viewer"

class UserUpdate(BaseModel):
    role: Optional[str] = None
    password: Optional[str] = None

class InventoryItem(BaseModel):
    item_name: str
    quantity: int
    group_name: Optional[str] = None
    custom_fields: Optional[Dict[str, Any]] = {}

class InventoryItemUpdate(BaseModel):
    item_name: Optional[str] = None
    quantity: Optional[int] = None
    group_name: Optional[str] = None
    custom_fields: Optional[Dict[str, Any]] = None

class GroupCreate(BaseModel):
    group_name: str
    description: Optional[str] = None

class GroupRename(BaseModel):
    new_name: str

class PriceUpdate(BaseModel):
    price: float
    supplier: Optional[str] = "default"

class PasswordChange(BaseModel):
    old_password: str
    new_password: str

class SearchQuery(BaseModel):
    search_term: str
    search_type: Optional[str] = "contains"  # starts_with, contains, exact

# New Pydantic Models for Inventory Tracking Features

class LocationCreate(BaseModel):
    name: str
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    zip_code: Optional[str] = None
    country: Optional[str] = "USA"
    location_type: Optional[str] = "warehouse"
    capacity: Optional[int] = None
    current_utilization: Optional[int] = 0
    manager_name: Optional[str] = None
    contact_phone: Optional[str] = None
    contact_email: Optional[str] = None
    is_active: Optional[bool] = True
    notes: Optional[str] = None

class LocationUpdate(BaseModel):
    name: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    zip_code: Optional[str] = None
    country: Optional[str] = None
    location_type: Optional[str] = None
    capacity: Optional[int] = None
    current_utilization: Optional[int] = None
    manager_name: Optional[str] = None
    contact_phone: Optional[str] = None
    contact_email: Optional[str] = None
    is_active: Optional[bool] = None
    notes: Optional[str] = None

class ItemLocationCreate(BaseModel):
    item_name: str
    location_id: int
    quantity: int
    aisle: Optional[str] = None
    shelf: Optional[str] = None
    bin: Optional[str] = None
    notes: Optional[str] = None

class BatchCreate(BaseModel):
    batch_number: str
    item_name: str
    location_id: Optional[int] = None
    quantity: int
    manufacturing_date: Optional[str] = None
    expiry_date: Optional[str] = None
    received_date: Optional[str] = None
    supplier_id: Optional[int] = None
    cost_per_unit: Optional[float] = None
    status: Optional[str] = "active"
    notes: Optional[str] = None

class BatchUpdate(BaseModel):
    quantity: Optional[int] = None
    location_id: Optional[int] = None
    status: Optional[str] = None
    notes: Optional[str] = None

class StockAdjustmentCreate(BaseModel):
    item_name: str
    location_id: Optional[int] = None
    batch_id: Optional[int] = None
    adjustment_type: str  # increase or decrease
    quantity: int
    reason: str
    reason_notes: Optional[str] = None
    adjusted_by: str
    approved_by: Optional[str] = None
    reference_number: Optional[str] = None

class AlertUpdate(BaseModel):
    is_read: Optional[bool] = None
    is_resolved: Optional[bool] = None
    resolved_by: Optional[str] = None

# ============================================================================
# Authentication
# ============================================================================

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def get_current_user(token: str = Depends(oauth2_scheme)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        role: str = payload.get("role")
        if username is None:
            raise credentials_exception
        token_data = TokenData(username=username, role=role)
    except JWTError:
        raise credentials_exception

    user = User(username=token_data.username, role=token_data.role)
    if user is None:
        raise credentials_exception
    return user

def get_admin_user(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

def get_editor_user(current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "editor"]:
        raise HTTPException(status_code=403, detail="Editor or admin access required")
    return current_user

def get_admin_or_editor(current_user: User = Depends(get_current_user)):
    """Alias for get_editor_user for clarity"""
    if current_user.role not in ["admin", "editor"]:
        raise HTTPException(status_code=403, detail="Editor or admin access required")
    return current_user

# ============================================================================
# Authentication Endpoints
# ============================================================================

@app.post("/token", response_model=Token)
async def login(form_data: OAuth2PasswordRequestForm = Depends()):
    """Login endpoint - returns JWT token"""
    logging.info(f"Login attempt for user: {form_data.username}")

    role = user_service.authenticate(form_data.username, form_data.password)
    if not role:
        logging.warning(f"Failed login attempt for user: {form_data.username}")
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )

    access_token = create_access_token(data={"sub": form_data.username, "role": role})
    logging.info(f"User {form_data.username} logged in successfully with role: {role}")

    return {"access_token": access_token, "token_type": "bearer"}

@app.get("/users/me", response_model=User)
async def read_users_me(current_user: User = Depends(get_current_user)):
    """Get current user info"""
    return current_user

# ============================================================================
# User Management Endpoints
# ============================================================================

@app.get("/users")
async def get_users(current_user: User = Depends(get_admin_user)):
    """Get all users (admin only)"""
    users = user_service.get_all_users()
    return {"users": users}

@app.post("/users")
async def create_user(user: UserCreate, current_user: User = Depends(get_admin_user)):
    """Create new user (admin only)"""
    try:
        result = user_service.create_user(user.username, user.password, user.role)
        return {"message": "User created successfully", "user": result}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.put("/users/{username}")
async def update_user(username: str, user_update: UserUpdate, current_user: User = Depends(get_admin_user)):
    """Update user (admin only)"""
    try:
        if user_update.password:
            user_service.update_password(username, user_update.password)
        if user_update.role:
            result = user_service.update_user(username, role=user_update.role)
            return {"message": "User updated successfully", "user": result}
        return {"message": "No updates provided"}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.delete("/users/{username}")
async def delete_user(username: str, current_user: User = Depends(get_admin_user)):
    """Delete user (admin only)"""
    try:
        success = user_service.delete_user(username)
        if success:
            return {"message": f"User '{username}' deleted successfully"}
        raise HTTPException(status_code=404, detail="User not found")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/users/me/change-password")
async def change_my_password(
    password_change: PasswordChange,
    current_user: User = Depends(get_current_user)
):
    """Change current user's password"""
    try:
        user_service.change_password(
            current_user.username,
            password_change.old_password,
            password_change.new_password
        )
        return {"message": "Password changed successfully"}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

# ============================================================================
# Inventory Endpoints
# ============================================================================

@app.get("/inventory")
async def get_inventory(current_user: User = Depends(get_current_user)):
    """Get all inventory items"""
    items = inventory_service.get_inventory()
    # Convert Item objects to dictionaries
    items_dict = []
    for item in items:
        items_dict.append({
            "item_name": item.item_name,
            "quantity": item.quantity,
            "group_name": item.group_name,
            "custom_fields": item.custom_fields or {}
        })
    return items_dict

@app.get("/inventory/{item_name}")
async def get_item(item_name: str, current_user: User = Depends(get_current_user)):
    """Get specific item"""
    item = inventory_service.get_item(item_name)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    return {
        "item_name": item.item_name,
        "quantity": item.quantity,
        "group_name": item.group_name,
        "custom_fields": item.custom_fields or {}
    }

@app.get("/inventory/check-duplicate/{item_name}")
async def check_duplicate_item(item_name: str, current_user: User = Depends(get_current_user)):
    """Check for similar/duplicate items"""
    try:
        with db_connection.get_cursor() as cursor:
            # Check exact match
            cursor.execute("SELECT 1 FROM items WHERE item_name = ?", (item_name,))
            exact_match = cursor.fetchone()

            # Check similar names (fuzzy match)
            similar_items = []
            cursor.execute("SELECT item_name FROM items")
            all_items = cursor.fetchall()

            item_lower = item_name.lower()
            for row in all_items:
                existing_name = row['item_name']
                existing_lower = existing_name.lower()

                # Check various similarity conditions
                if (item_lower in existing_lower or existing_lower in item_lower or
                    item_lower.replace(' ', '') == existing_lower.replace(' ', '') or
                    item_lower.replace('-', '') == existing_lower.replace('-', '')):
                    similar_items.append(existing_name)

            return {
                "exists": exact_match is not None,
                "similar_items": similar_items[:5]  # Limit to 5 suggestions
            }
    except Exception as e:
        logging.error(f"Error checking duplicates: {e}")
        raise HTTPException(status_code=500, detail="Error checking for duplicates")

@app.post("/inventory")
async def add_inventory_item(item: InventoryItem, current_user: User = Depends(get_editor_user)):
    """Add new inventory item"""
    success = inventory_service.add_item(
        item.item_name,
        item.quantity,
        item.group_name,
        item.custom_fields
    )
    if success:
        return {"message": "Item added successfully", "item_name": item.item_name}
    raise HTTPException(status_code=400, detail="Failed to add item")

@app.put("/inventory/{item_name}")
async def update_inventory_item(
    item_name: str,
    item_update: InventoryItemUpdate,
    current_user: User = Depends(get_editor_user)
):
    """Update inventory item"""
    # Get existing item
    existing_item = inventory_service.get_item(item_name)
    if not existing_item:
        raise HTTPException(status_code=404, detail="Item not found")

    # Delete and recreate if name changed
    if item_update.item_name and item_update.item_name != item_name:
        inventory_service.delete_item(item_name)
        success = inventory_service.add_item(
            item_update.item_name,
            item_update.quantity if item_update.quantity is not None else existing_item.quantity,
            item_update.group_name if item_update.group_name is not None else existing_item.group_name,
            item_update.custom_fields if item_update.custom_fields is not None else existing_item.custom_fields
        )
    else:
        # Update existing
        inventory_service.delete_item(item_name)
        success = inventory_service.add_item(
            item_name,
            item_update.quantity if item_update.quantity is not None else existing_item.quantity,
            item_update.group_name if item_update.group_name is not None else existing_item.group_name,
            item_update.custom_fields if item_update.custom_fields is not None else existing_item.custom_fields
        )

    if success:
        return {"message": "Item updated successfully"}
    raise HTTPException(status_code=400, detail="Failed to update item")

@app.delete("/inventory/{item_name}")
async def delete_inventory_item(item_name: str, current_user: User = Depends(get_admin_user)):
    """Delete inventory item"""
    success = inventory_service.delete_item(item_name)
    if success:
        return {"message": f"Item '{item_name}' deleted successfully"}
    raise HTTPException(status_code=404, detail="Item not found")

class BulkDeleteRequest(BaseModel):
    item_names: List[str]

class BulkUpdateRequest(BaseModel):
    item_names: List[str]
    quantity: Optional[int] = None
    group_name: Optional[str] = None
    reorder_level: Optional[int] = None
    reorder_quantity: Optional[int] = None

@app.post("/inventory/bulk-delete")
async def bulk_delete_items(request: BulkDeleteRequest, current_user: User = Depends(get_admin_user)):
    """Bulk delete multiple items"""
    try:
        deleted = []
        failed = []

        for item_name in request.item_names:
            success = inventory_service.delete_item(item_name)
            if success:
                deleted.append(item_name)
            else:
                failed.append(item_name)

        return {
            "message": f"Deleted {len(deleted)} items",
            "deleted": deleted,
            "failed": failed
        }
    except Exception as e:
        logging.error(f"Error in bulk delete: {e}")
        raise HTTPException(status_code=500, detail="Error performing bulk delete")

@app.post("/inventory/bulk-update")
async def bulk_update_items(request: BulkUpdateRequest, current_user: User = Depends(get_admin_or_editor)):
    """Bulk update multiple items"""
    try:
        with db_connection.get_cursor() as cursor:
            updated = []
            failed = []

            for item_name in request.item_names:
                try:
                    # Check if item exists
                    cursor.execute("SELECT 1 FROM items WHERE item_name = ?", (item_name,))
                    if not cursor.fetchone():
                        failed.append({"item": item_name, "reason": "not found"})
                        continue

                    # Build update query dynamically
                    updates = []
                    params = []

                    if request.quantity is not None:
                        updates.append("quantity = ?")
                        params.append(request.quantity)

                    if request.group_name is not None:
                        updates.append("group_name = ?")
                        params.append(request.group_name)

                    if request.reorder_level is not None:
                        updates.append("reorder_level = ?")
                        params.append(request.reorder_level)

                    if request.reorder_quantity is not None:
                        updates.append("reorder_quantity = ?")
                        params.append(request.reorder_quantity)

                    if updates:
                        params.append(item_name)
                        cursor.execute(f"""
                            UPDATE items
                            SET {', '.join(updates)}
                            WHERE item_name = ?
                        """, params)

                        # Add to history
                        cursor.execute("""
                            INSERT INTO history (action, item_name, user_name)
                            VALUES (?, ?, ?)
                        """, ("bulk_updated", item_name, current_user.username))

                        updated.append(item_name)

                except Exception as e:
                    failed.append({"item": item_name, "reason": str(e)})

            return {
                "message": f"Updated {len(updated)} items",
                "updated": updated,
                "failed": failed
            }
    except Exception as e:
        logging.error(f"Error in bulk update: {e}")
        raise HTTPException(status_code=500, detail="Error performing bulk update")

@app.get("/inventory/{item_name}/history")
async def get_item_history(item_name: str, current_user: User = Depends(get_current_user)):
    """Get item history"""
    history = inventory_service.get_item_history(item_name)
    history_list = []
    for entry in history:
        history_list.append({
            "action": entry.action,
            "item_name": entry.item_name,
            "quantity": entry.quantity,
            "group_name": entry.group_name,
            "timestamp": entry.timestamp
        })
    return {"history": history_list}

@app.post("/inventory/search")
async def search_inventory(
    search_query: SearchQuery,
    current_user: User = Depends(get_current_user)
):
    """Advanced search for inventory items"""
    items = inventory_service.search_items(
        search_query.search_term,
        search_query.search_type
    )
    items_dict = []
    for item in items:
        items_dict.append({
            "item_name": item.item_name,
            "quantity": item.quantity,
            "group_name": item.group_name,
            "custom_fields": item.custom_fields or {}
        })
    return {"items": items_dict, "count": len(items_dict)}

# ============================================================================
# Groups Endpoints
# ============================================================================

@app.get("/groups")
async def get_groups(current_user: User = Depends(get_current_user)):
    """Get all groups"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("SELECT group_name, description, created_at FROM groups ORDER BY group_name")
            groups = []
            for row in cursor.fetchall():
                groups.append({
                    "group_name": row['group_name'],
                    "description": row['description'],
                    "created_at": row['created_at']
                })
            return {"groups": groups}
    except Exception as e:
        logging.error(f"Error fetching groups: {e}")
        raise HTTPException(status_code=500, detail="Error fetching groups")

@app.post("/groups")
async def create_group(group: GroupCreate, current_user: User = Depends(get_editor_user)):
    """Create new group"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute(
                "INSERT INTO groups (group_name, description) VALUES (?, ?)",
                (group.group_name, group.description)
            )
            return {"message": "Group created successfully", "group_name": group.group_name}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Group already exists")
    except Exception as e:
        logging.error(f"Error creating group: {e}")
        raise HTTPException(status_code=500, detail="Error creating group")

@app.put("/groups/{old_name}")
async def rename_group(old_name: str, group_rename: GroupRename, current_user: User = Depends(get_editor_user)):
    """Rename a group"""
    new_name = group_rename.new_name
    try:
        with db_connection.get_cursor() as cursor:
            # Check if old group exists
            cursor.execute("SELECT 1 FROM groups WHERE group_name = ?", (old_name,))
            if not cursor.fetchone():
                raise HTTPException(status_code=404, detail=f"Group '{old_name}' not found")

            # Check if new name already exists
            if old_name != new_name:
                cursor.execute("SELECT 1 FROM groups WHERE group_name = ?", (new_name,))
                if cursor.fetchone():
                    raise HTTPException(status_code=400, detail=f"Group '{new_name}' already exists")

            # Update group name
            cursor.execute("UPDATE groups SET group_name = ? WHERE group_name = ?", (new_name, old_name))

            # Update items with this group
            cursor.execute("UPDATE items SET group_name = ? WHERE group_name = ?", (new_name, old_name))

            return {"message": f"Group renamed from '{old_name}' to '{new_name}'"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error renaming group: {e}")
        raise HTTPException(status_code=500, detail="Error renaming group")

@app.delete("/groups/{group_name}")
async def delete_group(group_name: str, current_user: User = Depends(get_admin_user)):
    """Delete a group"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("DELETE FROM groups WHERE group_name = ?", (group_name,))
            if cursor.rowcount > 0:
                # Set group_name to NULL for items in this group
                cursor.execute("UPDATE items SET group_name = NULL WHERE group_name = ?", (group_name,))
                return {"message": f"Group '{group_name}' deleted successfully"}
            raise HTTPException(status_code=404, detail="Group not found")
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting group: {e}")
        raise HTTPException(status_code=500, detail="Error deleting group")

# ============================================================================
# Prices Endpoints
# ============================================================================

@app.get("/prices")
async def get_all_prices(current_user: User = Depends(get_current_user)):
    """Get all prices"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("""
                SELECT item_name, price, supplier, date_updated, is_unit_price
                FROM prices
                ORDER BY item_name, supplier
            """)
            prices = []
            for row in cursor.fetchall():
                prices.append({
                    "item_name": row['item_name'],
                    "price": row['price'],
                    "supplier": row['supplier'],
                    "date_updated": row['date_updated'],
                    "is_unit_price": bool(row['is_unit_price'])
                })
            return {"prices": prices}
    except Exception as e:
        logging.error(f"Error fetching prices: {e}")
        raise HTTPException(status_code=500, detail="Error fetching prices")

@app.get("/prices/{item_name}")
async def get_item_price(item_name: str, current_user: User = Depends(get_current_user)):
    """Get price for specific item"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("""
                SELECT item_name, price, supplier, date_updated, is_unit_price
                FROM prices
                WHERE item_name = ?
                ORDER BY price ASC
            """, (item_name,))
            prices = []
            for row in cursor.fetchall():
                prices.append({
                    "item_name": row['item_name'],
                    "price": row['price'],
                    "supplier": row['supplier'],
                    "date_updated": row['date_updated'],
                    "is_unit_price": bool(row['is_unit_price'])
                })
            return {"prices": prices}
    except Exception as e:
        logging.error(f"Error fetching price: {e}")
        raise HTTPException(status_code=500, detail="Error fetching price")

@app.put("/prices/{item_name}")
async def update_price(
    item_name: str,
    price_update: PriceUpdate,
    current_user: User = Depends(get_editor_user)
):
    """Update item price"""
    try:
        with db_connection.get_cursor() as cursor:
            # Check if item exists
            cursor.execute("SELECT 1 FROM items WHERE item_name = ?", (item_name,))
            if not cursor.fetchone():
                raise HTTPException(status_code=404, detail="Item not found")

            # Insert or update price
            cursor.execute("SELECT 1 FROM prices WHERE item_name = ? AND supplier = ?", (item_name, price_update.supplier))
            if cursor.fetchone():
                # Update existing
                cursor.execute("""
                    UPDATE prices
                    SET price = ?, date_updated = datetime('now'), is_unit_price = 1
                    WHERE item_name = ? AND supplier = ?
                """, (price_update.price, item_name, price_update.supplier))
            else:
                # Insert new
                cursor.execute("""
                    INSERT INTO prices (item_name, price, supplier, date_updated, is_unit_price)
                    VALUES (?, ?, ?, datetime('now'), 1)
                """, (item_name, price_update.price, price_update.supplier))

            return {"message": "Price updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error updating price: {e}")
        raise HTTPException(status_code=500, detail="Error updating price")

@app.delete("/prices/{item_name}")
async def delete_price(item_name: str, supplier: str = "default", current_user: User = Depends(get_admin_user)):
    """Delete price entry"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("DELETE FROM prices WHERE item_name = ? AND supplier = ?", (item_name, supplier))
            if cursor.rowcount > 0:
                return {"message": "Price deleted successfully"}
            raise HTTPException(status_code=404, detail="Price not found")
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting price: {e}")
        raise HTTPException(status_code=500, detail="Error deleting price")

@app.get("/prices/{item_name}/cheapest")
async def get_cheapest_price(item_name: str, current_user: User = Depends(get_current_user)):
    """Get the cheapest price for an item across all suppliers"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("""
                SELECT item_name, price, supplier, date_updated
                FROM prices
                WHERE item_name = ?
                ORDER BY price ASC
                LIMIT 1
            """, (item_name,))
            result = cursor.fetchone()
            if result:
                return {
                    "item_name": result['item_name'],
                    "price": result['price'],
                    "supplier": result['supplier'],
                    "date_updated": result['date_updated']
                }
            raise HTTPException(status_code=404, detail="No prices found for this item")
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error fetching cheapest price: {e}")
        raise HTTPException(status_code=500, detail="Error fetching cheapest price")

@app.get("/prices/{item_name}/history")
async def get_price_history(item_name: str, current_user: User = Depends(get_current_user)):
    """Get price history for a specific item"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("""
                SELECT price, supplier, timestamp
                FROM price_history
                WHERE item_name = ?
                ORDER BY timestamp DESC
                LIMIT 100
            """, (item_name,))
            history = []
            for row in cursor.fetchall():
                history.append({
                    "price": row['price'],
                    "supplier": row['supplier'],
                    "date": row['timestamp']
                })
            return {"history": history, "item_name": item_name}
    except Exception as e:
        logging.error(f"Error fetching price history: {e}")
        raise HTTPException(status_code=500, detail="Error fetching price history")

@app.post("/prices/{item_name}")
async def add_price(
    item_name: str,
    price_update: PriceUpdate,
    current_user: User = Depends(get_editor_user)
):
    """Add a new price entry (alias for update)"""
    return await update_price(item_name, price_update, current_user)

@app.get("/prices/compare/all")
async def compare_all_prices(current_user: User = Depends(get_current_user)):
    """Get price comparison across all items and suppliers"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("""
                SELECT
                    item_name,
                    supplier,
                    price,
                    date_updated,
                    RANK() OVER (PARTITION BY item_name ORDER BY price ASC) as price_rank
                FROM prices
                ORDER BY item_name, price ASC
            """)
            comparison = {}
            for row in cursor.fetchall():
                item = row['item_name']
                if item not in comparison:
                    comparison[item] = {
                        'cheapest': None,
                        'most_expensive': None,
                        'suppliers': []
                    }

                supplier_data = {
                    'supplier': row['supplier'],
                    'price': row['price'],
                    'date_updated': row['date_updated'],
                    'is_cheapest': row['price_rank'] == 1
                }
                comparison[item]['suppliers'].append(supplier_data)

                if comparison[item]['cheapest'] is None or row['price'] < comparison[item]['cheapest']['price']:
                    comparison[item]['cheapest'] = supplier_data
                if comparison[item]['most_expensive'] is None or row['price'] > comparison[item]['most_expensive']['price']:
                    comparison[item]['most_expensive'] = supplier_data

            return {"comparison": comparison}
    except Exception as e:
        logging.error(f"Error comparing prices: {e}")
        raise HTTPException(status_code=500, detail="Error comparing prices")

# ============================================================================
# Supplier Management Endpoints
# ============================================================================

class SupplierCreate(BaseModel):
    name: str
    contact_person: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    zip_code: Optional[str] = None
    country: Optional[str] = "USA"
    website: Optional[str] = None
    notes: Optional[str] = None
    rating: Optional[int] = None
    is_active: Optional[bool] = True

class SupplierUpdate(BaseModel):
    name: Optional[str] = None
    contact_person: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    zip_code: Optional[str] = None
    country: Optional[str] = None
    website: Optional[str] = None
    notes: Optional[str] = None
    rating: Optional[int] = None
    is_active: Optional[bool] = None

@app.get("/suppliers")
async def get_all_suppliers(
    active_only: bool = False,
    current_user: User = Depends(get_current_user)
):
    """Get all suppliers"""
    try:
        with db_connection.get_cursor() as cursor:
            if active_only:
                cursor.execute("""
                    SELECT * FROM suppliers WHERE is_active = 1 ORDER BY name
                """)
            else:
                cursor.execute("""
                    SELECT * FROM suppliers ORDER BY name
                """)

            suppliers = []
            for row in cursor.fetchall():
                suppliers.append({
                    "id": row['id'],
                    "name": row['name'],
                    "contact_person": row['contact_person'],
                    "email": row['email'],
                    "phone": row['phone'],
                    "address": row['address'],
                    "city": row['city'],
                    "state": row['state'],
                    "zip_code": row['zip_code'],
                    "country": row['country'],
                    "website": row['website'],
                    "notes": row['notes'],
                    "rating": row['rating'],
                    "is_active": bool(row['is_active']),
                    "created_at": row['created_at'],
                    "updated_at": row['updated_at']
                })
            return {"suppliers": suppliers}
    except Exception as e:
        logging.error(f"Error fetching suppliers: {e}")
        raise HTTPException(status_code=500, detail="Error fetching suppliers")

@app.get("/suppliers/{supplier_id}")
async def get_supplier(supplier_id: int, current_user: User = Depends(get_current_user)):
    """Get a specific supplier by ID"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("SELECT * FROM suppliers WHERE id = ?", (supplier_id,))
            row = cursor.fetchone()
            if row:
                return {
                    "id": row['id'],
                    "name": row['name'],
                    "contact_person": row['contact_person'],
                    "email": row['email'],
                    "phone": row['phone'],
                    "address": row['address'],
                    "city": row['city'],
                    "state": row['state'],
                    "zip_code": row['zip_code'],
                    "country": row['country'],
                    "website": row['website'],
                    "notes": row['notes'],
                    "rating": row['rating'],
                    "is_active": bool(row['is_active']),
                    "created_at": row['created_at'],
                    "updated_at": row['updated_at']
                }
            raise HTTPException(status_code=404, detail="Supplier not found")
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error fetching supplier: {e}")
        raise HTTPException(status_code=500, detail="Error fetching supplier")

@app.post("/suppliers")
async def create_supplier(
    supplier: SupplierCreate,
    current_user: User = Depends(get_editor_user)
):
    """Create a new supplier"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("""
                INSERT INTO suppliers (
                    name, contact_person, email, phone, address, city, state,
                    zip_code, country, website, notes, rating, is_active
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                supplier.name, supplier.contact_person, supplier.email, supplier.phone,
                supplier.address, supplier.city, supplier.state, supplier.zip_code,
                supplier.country, supplier.website, supplier.notes, supplier.rating,
                1 if supplier.is_active else 0
            ))
            supplier_id = cursor.lastrowid
            return {
                "message": "Supplier created successfully",
                "id": supplier_id,
                "name": supplier.name
            }
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Supplier with this name already exists")
    except Exception as e:
        logging.error(f"Error creating supplier: {e}")
        raise HTTPException(status_code=500, detail="Error creating supplier")

@app.put("/suppliers/{supplier_id}")
async def update_supplier(
    supplier_id: int,
    supplier_update: SupplierUpdate,
    current_user: User = Depends(get_editor_user)
):
    """Update supplier information"""
    try:
        with db_connection.get_cursor() as cursor:
            # Check if supplier exists
            cursor.execute("SELECT 1 FROM suppliers WHERE id = ?", (supplier_id,))
            if not cursor.fetchone():
                raise HTTPException(status_code=404, detail="Supplier not found")

            # Build update query dynamically
            updates = []
            params = []

            if supplier_update.name is not None:
                updates.append("name = ?")
                params.append(supplier_update.name)
            if supplier_update.contact_person is not None:
                updates.append("contact_person = ?")
                params.append(supplier_update.contact_person)
            if supplier_update.email is not None:
                updates.append("email = ?")
                params.append(supplier_update.email)
            if supplier_update.phone is not None:
                updates.append("phone = ?")
                params.append(supplier_update.phone)
            if supplier_update.address is not None:
                updates.append("address = ?")
                params.append(supplier_update.address)
            if supplier_update.city is not None:
                updates.append("city = ?")
                params.append(supplier_update.city)
            if supplier_update.state is not None:
                updates.append("state = ?")
                params.append(supplier_update.state)
            if supplier_update.zip_code is not None:
                updates.append("zip_code = ?")
                params.append(supplier_update.zip_code)
            if supplier_update.country is not None:
                updates.append("country = ?")
                params.append(supplier_update.country)
            if supplier_update.website is not None:
                updates.append("website = ?")
                params.append(supplier_update.website)
            if supplier_update.notes is not None:
                updates.append("notes = ?")
                params.append(supplier_update.notes)
            if supplier_update.rating is not None:
                updates.append("rating = ?")
                params.append(supplier_update.rating)
            if supplier_update.is_active is not None:
                updates.append("is_active = ?")
                params.append(1 if supplier_update.is_active else 0)

            if not updates:
                return {"message": "No updates provided"}

            updates.append("updated_at = datetime('now')")
            params.append(supplier_id)

            query = f"UPDATE suppliers SET {', '.join(updates)} WHERE id = ?"
            cursor.execute(query, params)

            return {"message": "Supplier updated successfully"}
    except HTTPException:
        raise
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Supplier name already exists")
    except Exception as e:
        logging.error(f"Error updating supplier: {e}")
        raise HTTPException(status_code=500, detail="Error updating supplier")

@app.delete("/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: int, current_user: User = Depends(get_admin_user)):
    """Delete a supplier"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("DELETE FROM suppliers WHERE id = ?", (supplier_id,))
            if cursor.rowcount > 0:
                return {"message": "Supplier deleted successfully"}
            raise HTTPException(status_code=404, detail="Supplier not found")
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting supplier: {e}")
        raise HTTPException(status_code=500, detail="Error deleting supplier")

@app.get("/suppliers/{supplier_id}/items")
async def get_supplier_items(supplier_id: int, current_user: User = Depends(get_current_user)):
    """Get all items supplied by a specific supplier"""
    try:
        with db_connection.get_cursor() as cursor:
            # First get supplier name
            cursor.execute("SELECT name FROM suppliers WHERE id = ?", (supplier_id,))
            supplier_row = cursor.fetchone()
            if not supplier_row:
                raise HTTPException(status_code=404, detail="Supplier not found")

            supplier_name = supplier_row['name']

            # Get all items with prices from this supplier
            cursor.execute("""
                SELECT DISTINCT p.item_name, p.price, p.date_updated, i.quantity
                FROM prices p
                LEFT JOIN items i ON p.item_name = i.item_name
                WHERE p.supplier = ?
                ORDER BY p.item_name
            """, (supplier_name,))

            items = []
            for row in cursor.fetchall():
                items.append({
                    "item_name": row['item_name'],
                    "price": row['price'],
                    "quantity": row['quantity'],
                    "date_updated": row['date_updated']
                })

            return {
                "supplier_id": supplier_id,
                "supplier_name": supplier_name,
                "items": items,
                "total_items": len(items)
            }
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error fetching supplier items: {e}")
        raise HTTPException(status_code=500, detail="Error fetching supplier items")

@app.get("/suppliers/search/{name}")
async def search_suppliers(name: str, current_user: User = Depends(get_current_user)):
    """Search suppliers by name"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("""
                SELECT * FROM suppliers
                WHERE name LIKE ?
                ORDER BY name
            """, (f"%{name}%",))

            suppliers = []
            for row in cursor.fetchall():
                suppliers.append({
                    "id": row['id'],
                    "name": row['name'],
                    "contact_person": row['contact_person'],
                    "email": row['email'],
                    "phone": row['phone'],
                    "rating": row['rating'],
                    "is_active": bool(row['is_active'])
                })
            return {"suppliers": suppliers}
    except Exception as e:
        logging.error(f"Error searching suppliers: {e}")
        raise HTTPException(status_code=500, detail="Error searching suppliers")

# ============================================================================
# SUPPLIER-PRODUCT RELATIONSHIPS
# ============================================================================

class SupplierProduct(BaseModel):
    supplier_id: int
    item_name: str
    supplier_sku: Optional[str] = None
    unit_price: float
    minimum_order_quantity: Optional[int] = 1
    lead_time_days: Optional[int] = None
    is_available: Optional[bool] = True
    notes: Optional[str] = None

@app.get("/supplier-products/{supplier_id}")
async def get_supplier_products(supplier_id: int, current_user: User = Depends(get_current_user)):
    """Get all products offered by a supplier with prices"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("""
                SELECT sp.*, i.quantity as current_stock, i.group_name
                FROM supplier_products sp
                LEFT JOIN items i ON sp.item_name = i.item_name
                WHERE sp.supplier_id = ?
                ORDER BY sp.item_name
            """, (supplier_id,))

            products = []
            for row in cursor.fetchall():
                products.append({
                    "id": row['id'],
                    "item_name": row['item_name'],
                    "supplier_sku": row['supplier_sku'],
                    "unit_price": row['unit_price'],
                    "minimum_order_quantity": row['minimum_order_quantity'],
                    "lead_time_days": row['lead_time_days'],
                    "is_available": bool(row['is_available']),
                    "current_stock": row['current_stock'],
                    "group_name": row['group_name'],
                    "last_price_update": row['last_price_update'],
                    "notes": row['notes']
                })
            return products
    except Exception as e:
        logging.error(f"Error fetching supplier products: {e}")
        raise HTTPException(status_code=500, detail="Error fetching supplier products")

@app.get("/item-suppliers/{item_name}")
async def get_item_suppliers(item_name: str, current_user: User = Depends(get_current_user)):
    """Get all suppliers for a specific item with their prices"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("""
                SELECT sp.*, s.name as supplier_name, s.rating, s.lead_time_days as supplier_lead_time,
                       s.email, s.phone, s.is_active as supplier_active
                FROM supplier_products sp
                JOIN suppliers s ON sp.supplier_id = s.id
                WHERE sp.item_name = ?
                ORDER BY sp.unit_price ASC
            """, (item_name,))

            suppliers = []
            for row in cursor.fetchall():
                suppliers.append({
                    "id": row['id'],
                    "supplier_id": row['supplier_id'],
                    "supplier_name": row['supplier_name'],
                    "supplier_sku": row['supplier_sku'],
                    "unit_price": row['unit_price'],
                    "minimum_order_quantity": row['minimum_order_quantity'],
                    "lead_time_days": row['lead_time_days'] or row['supplier_lead_time'],
                    "is_available": bool(row['is_available']),
                    "rating": row['rating'],
                    "email": row['email'],
                    "phone": row['phone'],
                    "supplier_active": bool(row['supplier_active']),
                    "last_price_update": row['last_price_update'],
                    "notes": row['notes']
                })
            return suppliers
    except Exception as e:
        logging.error(f"Error fetching item suppliers: {e}")
        raise HTTPException(status_code=500, detail="Error fetching item suppliers")

@app.post("/supplier-products")
async def create_supplier_product(product: SupplierProduct, current_user: User = Depends(get_admin_or_editor)):
    """Add a product to a supplier's catalog"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("""
                INSERT INTO supplier_products
                (supplier_id, item_name, supplier_sku, unit_price, minimum_order_quantity,
                 lead_time_days, is_available, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (product.supplier_id, product.item_name, product.supplier_sku, product.unit_price,
                  product.minimum_order_quantity, product.lead_time_days,
                  1 if product.is_available else 0, product.notes))

            return {"message": "Supplier product added successfully", "id": cursor.lastrowid}
    except Exception as e:
        logging.error(f"Error creating supplier product: {e}")
        raise HTTPException(status_code=500, detail="Error creating supplier product")

@app.put("/supplier-products/{id}")
async def update_supplier_product(id: int, product: SupplierProduct, current_user: User = Depends(get_admin_or_editor)):
    """Update supplier product information"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("""
                UPDATE supplier_products
                SET supplier_sku = ?, unit_price = ?, minimum_order_quantity = ?,
                    lead_time_days = ?, is_available = ?, notes = ?,
                    last_price_update = datetime('now'), updated_at = datetime('now')
                WHERE id = ?
            """, (product.supplier_sku, product.unit_price, product.minimum_order_quantity,
                  product.lead_time_days, 1 if product.is_available else 0, product.notes, id))

            return {"message": "Supplier product updated successfully"}
    except Exception as e:
        logging.error(f"Error updating supplier product: {e}")
        raise HTTPException(status_code=500, detail="Error updating supplier product")

@app.delete("/supplier-products/{id}")
async def delete_supplier_product(id: int, current_user: User = Depends(get_admin_user)):
    """Remove a product from supplier's catalog"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("DELETE FROM supplier_products WHERE id = ?", (id,))
            return {"message": "Supplier product deleted successfully"}
    except Exception as e:
        logging.error(f"Error deleting supplier product: {e}")
        raise HTTPException(status_code=500, detail="Error deleting supplier product")

@app.get("/best-price/{item_name}")
async def get_best_price(item_name: str, location_id: Optional[int] = None, current_user: User = Depends(get_current_user)):
    """Find best price for an item, optionally considering location proximity"""
    try:
        with db_connection.get_cursor() as cursor:
            if location_id:
                # Consider shipping costs from supplier location
                cursor.execute("""
                    SELECT sp.*, s.name as supplier_name, s.rating,
                           sl.shipping_cost, sl.distance_km, sl.estimated_delivery_days
                    FROM supplier_products sp
                    JOIN suppliers s ON sp.supplier_id = s.id
                    LEFT JOIN supplier_locations sl ON s.id = sl.supplier_id AND sl.location_id = ?
                    WHERE sp.item_name = ? AND sp.is_available = 1 AND s.is_active = 1
                    ORDER BY (sp.unit_price + COALESCE(sl.shipping_cost, 0)) ASC
                    LIMIT 1
                """, (location_id, item_name))
            else:
                cursor.execute("""
                    SELECT sp.*, s.name as supplier_name, s.rating
                    FROM supplier_products sp
                    JOIN suppliers s ON sp.supplier_id = s.id
                    WHERE sp.item_name = ? AND sp.is_available = 1 AND s.is_active = 1
                    ORDER BY sp.unit_price ASC
                    LIMIT 1
                """, (item_name,))

            result = cursor.fetchone()
            if not result:
                raise HTTPException(status_code=404, detail="No suppliers found for this item")

            response = {
                "item_name": item_name,
                "supplier_id": result['supplier_id'],
                "supplier_name": result['supplier_name'],
                "unit_price": result['unit_price'],
                "rating": result['rating'],
                "lead_time_days": result['lead_time_days'],
                "minimum_order_quantity": result['minimum_order_quantity']
            }

            if location_id and result.get('shipping_cost') is not None:
                response.update({
                    "shipping_cost": result['shipping_cost'],
                    "total_cost": result['unit_price'] + result['shipping_cost'],
                    "distance_km": result['distance_km'],
                    "estimated_delivery_days": result['estimated_delivery_days']
                })

            return response
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error finding best price: {e}")
        raise HTTPException(status_code=500, detail="Error finding best price")

# ============================================================================
# SUPPLIER-LOCATION RELATIONSHIPS
# ============================================================================

class SupplierLocation(BaseModel):
    supplier_id: int
    location_id: int
    distance_km: Optional[float] = None
    estimated_delivery_days: Optional[int] = None
    shipping_cost: Optional[float] = 0
    is_preferred: Optional[bool] = False
    notes: Optional[str] = None

@app.get("/supplier-locations/{supplier_id}")
async def get_supplier_locations(supplier_id: int, current_user: User = Depends(get_current_user)):
    """Get all locations a supplier can deliver to"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("""
                SELECT sl.*, l.name as location_name, l.city, l.state, l.location_type
                FROM supplier_locations sl
                JOIN locations l ON sl.location_id = l.id
                WHERE sl.supplier_id = ?
                ORDER BY sl.distance_km ASC
            """, (supplier_id,))

            locations = []
            for row in cursor.fetchall():
                locations.append({
                    "id": row['id'],
                    "location_id": row['location_id'],
                    "location_name": row['location_name'],
                    "city": row['city'],
                    "state": row['state'],
                    "location_type": row['location_type'],
                    "distance_km": row['distance_km'],
                    "estimated_delivery_days": row['estimated_delivery_days'],
                    "shipping_cost": row['shipping_cost'],
                    "is_preferred": bool(row['is_preferred']),
                    "notes": row['notes']
                })
            return locations
    except Exception as e:
        logging.error(f"Error fetching supplier locations: {e}")
        raise HTTPException(status_code=500, detail="Error fetching supplier locations")

@app.get("/location-suppliers/{location_id}")
async def get_location_suppliers(location_id: int, current_user: User = Depends(get_current_user)):
    """Get all suppliers that deliver to a location"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("""
                SELECT sl.*, s.name as supplier_name, s.rating, s.email, s.phone,
                       s.is_active, s.payment_terms
                FROM supplier_locations sl
                JOIN suppliers s ON sl.supplier_id = s.id
                WHERE sl.location_id = ?
                ORDER BY sl.distance_km ASC, s.rating DESC
            """, (location_id,))

            suppliers = []
            for row in cursor.fetchall():
                suppliers.append({
                    "id": row['id'],
                    "supplier_id": row['supplier_id'],
                    "supplier_name": row['supplier_name'],
                    "rating": row['rating'],
                    "email": row['email'],
                    "phone": row['phone'],
                    "is_active": bool(row['is_active']),
                    "distance_km": row['distance_km'],
                    "estimated_delivery_days": row['estimated_delivery_days'],
                    "shipping_cost": row['shipping_cost'],
                    "is_preferred": bool(row['is_preferred']),
                    "payment_terms": row['payment_terms'],
                    "notes": row['notes']
                })
            return suppliers
    except Exception as e:
        logging.error(f"Error fetching location suppliers: {e}")
        raise HTTPException(status_code=500, detail="Error fetching location suppliers")

@app.post("/supplier-locations")
async def create_supplier_location(sl: SupplierLocation, current_user: User = Depends(get_admin_or_editor)):
    """Link a supplier to a location with delivery details"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("""
                INSERT INTO supplier_locations
                (supplier_id, location_id, distance_km, estimated_delivery_days, shipping_cost, is_preferred, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (sl.supplier_id, sl.location_id, sl.distance_km, sl.estimated_delivery_days,
                  sl.shipping_cost, 1 if sl.is_preferred else 0, sl.notes))

            return {"message": "Supplier-location relationship created successfully", "id": cursor.lastrowid}
    except Exception as e:
        logging.error(f"Error creating supplier-location: {e}")
        raise HTTPException(status_code=500, detail="Error creating supplier-location relationship")

@app.put("/supplier-locations/{id}")
async def update_supplier_location(id: int, sl: SupplierLocation, current_user: User = Depends(get_admin_or_editor)):
    """Update supplier-location relationship"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("""
                UPDATE supplier_locations
                SET distance_km = ?, estimated_delivery_days = ?, shipping_cost = ?,
                    is_preferred = ?, notes = ?, updated_at = datetime('now')
                WHERE id = ?
            """, (sl.distance_km, sl.estimated_delivery_days, sl.shipping_cost,
                  1 if sl.is_preferred else 0, sl.notes, id))

            return {"message": "Supplier-location updated successfully"}
    except Exception as e:
        logging.error(f"Error updating supplier-location: {e}")
        raise HTTPException(status_code=500, detail="Error updating supplier-location")

@app.delete("/supplier-locations/{id}")
async def delete_supplier_location(id: int, current_user: User = Depends(get_admin_user)):
    """Remove supplier-location relationship"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("DELETE FROM supplier_locations WHERE id = ?", (id,))
            return {"message": "Supplier-location deleted successfully"}
    except Exception as e:
        logging.error(f"Error deleting supplier-location: {e}")
        raise HTTPException(status_code=500, detail="Error deleting supplier-location")

# ============================================================================
# LOCATIONS MANAGEMENT ENDPOINTS
# ============================================================================

@app.get("/locations")
async def get_all_locations(
    active_only: bool = False,
    current_user: User = Depends(get_current_user)
):
    """Get all locations"""
    try:
        with db_connection.get_cursor() as cursor:
            if active_only:
                cursor.execute("SELECT * FROM locations WHERE is_active = 1 ORDER BY name")
            else:
                cursor.execute("SELECT * FROM locations ORDER BY name")

            locations = []
            for row in cursor.fetchall():
                locations.append({
                    "id": row['id'],
                    "name": row['name'],
                    "address": row['address'],
                    "city": row['city'],
                    "state": row['state'],
                    "zip_code": row['zip_code'],
                    "country": row['country'],
                    "location_type": row['location_type'],
                    "capacity": row['capacity'],
                    "current_utilization": row['current_utilization'],
                    "manager_name": row['manager_name'],
                    "contact_phone": row['contact_phone'],
                    "contact_email": row['contact_email'],
                    "is_active": bool(row['is_active']),
                    "notes": row['notes'],
                    "created_at": row['created_at'],
                    "updated_at": row['updated_at']
                })
            return {"locations": locations}
    except Exception as e:
        logging.error(f"Error fetching locations: {e}")
        raise HTTPException(status_code=500, detail="Error fetching locations")

@app.get("/locations/{location_id}")
async def get_location(location_id: int, current_user: User = Depends(get_current_user)):
    """Get specific location"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("SELECT * FROM locations WHERE id = ?", (location_id,))
            row = cursor.fetchone()
            if row:
                return {
                    "id": row['id'],
                    "name": row['name'],
                    "address": row['address'],
                    "city": row['city'],
                    "state": row['state'],
                    "zip_code": row['zip_code'],
                    "country": row['country'],
                    "location_type": row['location_type'],
                    "capacity": row['capacity'],
                    "current_utilization": row['current_utilization'],
                    "manager_name": row['manager_name'],
                    "contact_phone": row['contact_phone'],
                    "contact_email": row['contact_email'],
                    "is_active": bool(row['is_active']),
                    "notes": row['notes'],
                    "created_at": row['created_at'],
                    "updated_at": row['updated_at']
                }
            raise HTTPException(status_code=404, detail="Location not found")
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error fetching location: {e}")
        raise HTTPException(status_code=500, detail="Error fetching location")

@app.post("/locations")
async def create_location(
    location: LocationCreate,
    current_user: User = Depends(get_editor_user)
):
    """Create new location"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("""
                INSERT INTO locations (
                    name, address, city, state, zip_code, country,
                    location_type, capacity, current_utilization, manager_name,
                    contact_phone, contact_email, is_active, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                location.name, location.address, location.city, location.state,
                location.zip_code, location.country, location.location_type,
                location.capacity, location.current_utilization, location.manager_name,
                location.contact_phone, location.contact_email,
                1 if location.is_active else 0, location.notes
            ))
            location_id = cursor.lastrowid
            return {"message": "Location created successfully", "id": location_id}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Location with this name already exists")
    except Exception as e:
        logging.error(f"Error creating location: {e}")
        raise HTTPException(status_code=500, detail="Error creating location")

@app.put("/locations/{location_id}")
async def update_location(
    location_id: int,
    location_update: LocationUpdate,
    current_user: User = Depends(get_editor_user)
):
    """Update location"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("SELECT 1 FROM locations WHERE id = ?", (location_id,))
            if not cursor.fetchone():
                raise HTTPException(status_code=404, detail="Location not found")

            updates = []
            params = []

            for field in ['name', 'address', 'city', 'state', 'zip_code', 'country',
                         'location_type', 'capacity', 'current_utilization', 'manager_name',
                         'contact_phone', 'contact_email', 'notes']:
                value = getattr(location_update, field, None)
                if value is not None:
                    updates.append(f"{field} = ?")
                    params.append(value)

            if location_update.is_active is not None:
                updates.append("is_active = ?")
                params.append(1 if location_update.is_active else 0)

            if not updates:
                return {"message": "No updates provided"}

            updates.append("updated_at = datetime('now')")
            params.append(location_id)

            query = f"UPDATE locations SET {', '.join(updates)} WHERE id = ?"
            cursor.execute(query, params)

            return {"message": "Location updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error updating location: {e}")
        raise HTTPException(status_code=500, detail="Error updating location")

@app.delete("/locations/{location_id}")
async def delete_location(location_id: int, current_user: User = Depends(get_admin_user)):
    """Delete location"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("DELETE FROM locations WHERE id = ?", (location_id,))
            if cursor.rowcount > 0:
                return {"message": "Location deleted successfully"}
            raise HTTPException(status_code=404, detail="Location not found")
    except Exception as e:
        logging.error(f"Error deleting location: {e}")
        raise HTTPException(status_code=500, detail="Error deleting location")

@app.get("/locations/{location_id}/items")
async def get_location_items(location_id: int, current_user: User = Depends(get_current_user)):
    """Get all items in a specific location"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("""
                SELECT il.*, i.quantity as total_quantity, i.group_name, i.custom_fields
                FROM item_locations il
                LEFT JOIN items i ON il.item_name = i.item_name
                WHERE il.location_id = ?
                ORDER BY il.item_name
            """, (location_id,))

            items = []
            for row in cursor.fetchall():
                items.append({
                    "id": row['id'],
                    "item_name": row['item_name'],
                    "quantity": row['quantity'],
                    "total_quantity": row['total_quantity'],
                    "aisle": row['aisle'],
                    "shelf": row['shelf'],
                    "bin": row['bin'],
                    "notes": row['notes'],
                    "last_counted": row['last_counted'],
                    "group_name": row['group_name']
                })
            return {"items": items, "total_items": len(items)}
    except Exception as e:
        logging.error(f"Error fetching location items: {e}")
        raise HTTPException(status_code=500, detail="Error fetching location items")

# ============================================================================
# ITEM LOCATIONS ENDPOINTS
# ============================================================================

@app.post("/item-locations")
async def assign_item_to_location(
    item_location: ItemLocationCreate,
    current_user: User = Depends(get_editor_user)
):
    """Assign item to location with quantity"""
    try:
        with db_connection.get_cursor() as cursor:
            # Check if item and location exist
            cursor.execute("SELECT 1 FROM items WHERE item_name = ?", (item_location.item_name,))
            if not cursor.fetchone():
                raise HTTPException(status_code=404, detail="Item not found")

            cursor.execute("SELECT 1 FROM locations WHERE id = ?", (item_location.location_id,))
            if not cursor.fetchone():
                raise HTTPException(status_code=404, detail="Location not found")

            cursor.execute("""
                INSERT INTO item_locations (item_name, location_id, quantity, aisle, shelf, bin, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                item_location.item_name, item_location.location_id, item_location.quantity,
                item_location.aisle, item_location.shelf, item_location.bin, item_location.notes
            ))

            return {"message": "Item assigned to location successfully"}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Item already assigned to this location")
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error assigning item to location: {e}")
        raise HTTPException(status_code=500, detail="Error assigning item to location")

@app.get("/items/{item_name}/locations")
async def get_item_locations(item_name: str, current_user: User = Depends(get_current_user)):
    """Get all locations where an item is stored"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("""
                SELECT il.*, l.name as location_name, l.location_type, l.city, l.state
                FROM item_locations il
                LEFT JOIN locations l ON il.location_id = l.id
                WHERE il.item_name = ?
                ORDER BY il.quantity DESC
            """, (item_name,))

            locations = []
            total_qty = 0
            for row in cursor.fetchall():
                locations.append({
                    "id": row['id'],
                    "location_id": row['location_id'],
                    "location_name": row['location_name'],
                    "location_type": row['location_type'],
                    "city": row['city'],
                    "state": row['state'],
                    "quantity": row['quantity'],
                    "aisle": row['aisle'],
                    "shelf": row['shelf'],
                    "bin": row['bin'],
                    "notes": row['notes'],
                    "last_counted": row['last_counted']
                })
                total_qty += row['quantity']

            return {"locations": locations, "total_locations": len(locations), "total_quantity": total_qty}
    except Exception as e:
        logging.error(f"Error fetching item locations: {e}")
        raise HTTPException(status_code=500, detail="Error fetching item locations")

# ============================================================================
# BATCHES MANAGEMENT ENDPOINTS
# ============================================================================

@app.get("/batches")
async def get_all_batches(
    status: Optional[str] = None,
    expiring_soon: bool = False,
    current_user: User = Depends(get_current_user)
):
    """Get all batches with optional filters"""
    try:
        with db_connection.get_cursor() as cursor:
            if expiring_soon:
                # Get batches expiring in next 30 days
                query = """
                    SELECT b.*, l.name as location_name, s.name as supplier_name
                    FROM batches b
                    LEFT JOIN locations l ON b.location_id = l.id
                    LEFT JOIN suppliers s ON b.supplier_id = s.id
                    WHERE b.expiry_date IS NOT NULL
                    AND b.expiry_date <= date('now', '+30 days')
                    AND b.status = 'active'
                    ORDER BY b.expiry_date ASC
                """
                cursor.execute(query)
            elif status:
                query = """
                    SELECT b.*, l.name as location_name, s.name as supplier_name
                    FROM batches b
                    LEFT JOIN locations l ON b.location_id = l.id
                    LEFT JOIN suppliers s ON b.supplier_id = s.id
                    WHERE b.status = ?
                    ORDER BY b.created_at DESC
                """
                cursor.execute(query, (status,))
            else:
                query = """
                    SELECT b.*, l.name as location_name, s.name as supplier_name
                    FROM batches b
                    LEFT JOIN locations l ON b.location_id = l.id
                    LEFT JOIN suppliers s ON b.supplier_id = s.id
                    ORDER BY b.created_at DESC
                """
                cursor.execute(query)

            batches = []
            for row in cursor.fetchall():
                batches.append({
                    "id": row['id'],
                    "batch_number": row['batch_number'],
                    "item_name": row['item_name'],
                    "location_id": row['location_id'],
                    "location_name": row['location_name'],
                    "quantity": row['quantity'],
                    "manufacturing_date": row['manufacturing_date'],
                    "expiry_date": row['expiry_date'],
                    "received_date": row['received_date'],
                    "supplier_id": row['supplier_id'],
                    "supplier_name": row['supplier_name'],
                    "cost_per_unit": row['cost_per_unit'],
                    "status": row['status'],
                    "notes": row['notes'],
                    "created_at": row['created_at'],
                    "updated_at": row['updated_at']
                })
            return {"batches": batches, "total": len(batches)}
    except Exception as e:
        logging.error(f"Error fetching batches: {e}")
        raise HTTPException(status_code=500, detail="Error fetching batches")

@app.post("/batches")
async def create_batch(
    batch: BatchCreate,
    current_user: User = Depends(get_editor_user)
):
    """Create new batch"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("""
                INSERT INTO batches (
                    batch_number, item_name, location_id, quantity, manufacturing_date,
                    expiry_date, received_date, supplier_id, cost_per_unit, status, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                batch.batch_number, batch.item_name, batch.location_id, batch.quantity,
                batch.manufacturing_date, batch.expiry_date, batch.received_date,
                batch.supplier_id, batch.cost_per_unit, batch.status, batch.notes
            ))
            batch_id = cursor.lastrowid

            # Check for expiry and create alert if needed
            if batch.expiry_date:
                from datetime import datetime as dt
                expiry = dt.strptime(batch.expiry_date, "%Y-%m-%d").date()
                today = dt.now().date()
                days_until_expiry = (expiry - today).days

                if days_until_expiry <= 30 and days_until_expiry > 0:
                    severity = 'critical' if days_until_expiry <= 7 else 'high' if days_until_expiry <= 14 else 'medium'
                    cursor.execute("""
                        INSERT INTO alerts (alert_type, severity, item_name, batch_id, message)
                        VALUES ('expiring_soon', ?, ?, ?, ?)
                    """, (
                        severity, batch.item_name, batch_id,
                        f"Batch {batch.batch_number} expires in {days_until_expiry} days"
                    ))

            return {"message": "Batch created successfully", "id": batch_id}
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=400, detail="Batch number already exists")
    except Exception as e:
        logging.error(f"Error creating batch: {e}")
        raise HTTPException(status_code=500, detail="Error creating batch")

@app.put("/batches/{batch_id}")
async def update_batch(
    batch_id: int,
    batch_update: BatchUpdate,
    current_user: User = Depends(get_editor_user)
):
    """Update batch"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("SELECT 1 FROM batches WHERE id = ?", (batch_id,))
            if not cursor.fetchone():
                raise HTTPException(status_code=404, detail="Batch not found")

            updates = []
            params = []

            if batch_update.quantity is not None:
                updates.append("quantity = ?")
                params.append(batch_update.quantity)
            if batch_update.location_id is not None:
                updates.append("location_id = ?")
                params.append(batch_update.location_id)
            if batch_update.status is not None:
                updates.append("status = ?")
                params.append(batch_update.status)
            if batch_update.notes is not None:
                updates.append("notes = ?")
                params.append(batch_update.notes)

            if not updates:
                return {"message": "No updates provided"}

            updates.append("updated_at = datetime('now')")
            params.append(batch_id)

            query = f"UPDATE batches SET {', '.join(updates)} WHERE id = ?"
            cursor.execute(query, params)

            return {"message": "Batch updated successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error updating batch: {e}")
        raise HTTPException(status_code=500, detail="Error updating batch")

@app.get("/items/{item_name}/batches")
async def get_item_batches(
    item_name: str,
    active_only: bool = True,
    current_user: User = Depends(get_current_user)
):
    """Get all batches for a specific item"""
    try:
        with db_connection.get_cursor() as cursor:
            if active_only:
                query = """
                    SELECT b.*, l.name as location_name, s.name as supplier_name
                    FROM batches b
                    LEFT JOIN locations l ON b.location_id = l.id
                    LEFT JOIN suppliers s ON b.supplier_id = s.id
                    WHERE b.item_name = ? AND b.status = 'active'
                    ORDER BY b.expiry_date ASC NULLS LAST
                """
            else:
                query = """
                    SELECT b.*, l.name as location_name, s.name as supplier_name
                    FROM batches b
                    LEFT JOIN locations l ON b.location_id = l.id
                    LEFT JOIN suppliers s ON b.supplier_id = s.id
                    WHERE b.item_name = ?
                    ORDER BY b.created_at DESC
                """

            cursor.execute(query, (item_name,))

            batches = []
            for row in cursor.fetchall():
                batches.append({
                    "id": row['id'],
                    "batch_number": row['batch_number'],
                    "location_id": row['location_id'],
                    "location_name": row['location_name'],
                    "quantity": row['quantity'],
                    "manufacturing_date": row['manufacturing_date'],
                    "expiry_date": row['expiry_date'],
                    "received_date": row['received_date'],
                    "supplier_id": row['supplier_id'],
                    "supplier_name": row['supplier_name'],
                    "cost_per_unit": row['cost_per_unit'],
                    "status": row['status'],
                    "notes": row['notes']
                })

            return {"batches": batches, "total": len(batches)}
    except Exception as e:
        logging.error(f"Error fetching item batches: {e}")
        raise HTTPException(status_code=500, detail="Error fetching item batches")

# ============================================================================
# STOCK ADJUSTMENTS ENDPOINTS
# ============================================================================

@app.get("/stock-adjustments")
async def get_stock_adjustments(
    item_name: Optional[str] = None,
    limit: int = 100,
    current_user: User = Depends(get_current_user)
):
    """Get stock adjustments with optional filters"""
    try:
        with db_connection.get_cursor() as cursor:
            if item_name:
                query = """
                    SELECT sa.*, l.name as location_name, b.batch_number
                    FROM stock_adjustments sa
                    LEFT JOIN locations l ON sa.location_id = l.id
                    LEFT JOIN batches b ON sa.batch_id = b.id
                    WHERE sa.item_name = ?
                    ORDER BY sa.adjustment_date DESC
                    LIMIT ?
                """
                cursor.execute(query, (item_name, limit))
            else:
                query = """
                    SELECT sa.*, l.name as location_name, b.batch_number
                    FROM stock_adjustments sa
                    LEFT JOIN locations l ON sa.location_id = l.id
                    LEFT JOIN batches b ON sa.batch_id = b.id
                    ORDER BY sa.adjustment_date DESC
                    LIMIT ?
                """
                cursor.execute(query, (limit,))

            adjustments = []
            for row in cursor.fetchall():
                adjustments.append({
                    "id": row['id'],
                    "item_name": row['item_name'],
                    "location_id": row['location_id'],
                    "location_name": row['location_name'],
                    "batch_id": row['batch_id'],
                    "batch_number": row['batch_number'],
                    "adjustment_type": row['adjustment_type'],
                    "quantity": row['quantity'],
                    "reason": row['reason'],
                    "reason_notes": row['reason_notes'],
                    "adjusted_by": row['adjusted_by'],
                    "approved_by": row['approved_by'],
                    "reference_number": row['reference_number'],
                    "adjustment_date": row['adjustment_date'],
                    "created_at": row['created_at']
                })

            return {"adjustments": adjustments, "total": len(adjustments)}
    except Exception as e:
        logging.error(f"Error fetching stock adjustments: {e}")
        raise HTTPException(status_code=500, detail="Error fetching stock adjustments")

@app.post("/stock-adjustments")
async def create_stock_adjustment(
    adjustment: StockAdjustmentCreate,
    current_user: User = Depends(get_editor_user)
):
    """Create stock adjustment and update inventory"""
    try:
        with db_connection.get_cursor() as cursor:
            # Insert adjustment record
            cursor.execute("""
                INSERT INTO stock_adjustments (
                    item_name, location_id, batch_id, adjustment_type, quantity,
                    reason, reason_notes, adjusted_by, approved_by, reference_number
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                adjustment.item_name, adjustment.location_id, adjustment.batch_id,
                adjustment.adjustment_type, adjustment.quantity, adjustment.reason,
                adjustment.reason_notes, adjustment.adjusted_by, adjustment.approved_by,
                adjustment.reference_number
            ))

            # Update item quantity in main items table
            if adjustment.adjustment_type == "increase":
                cursor.execute("""
                    UPDATE items SET quantity = quantity + ? WHERE item_name = ?
                """, (adjustment.quantity, adjustment.item_name))
            else:  # decrease
                cursor.execute("""
                    UPDATE items SET quantity = quantity - ? WHERE item_name = ?
                """, (adjustment.quantity, adjustment.item_name))

            # Update location quantity if specified
            if adjustment.location_id:
                if adjustment.adjustment_type == "increase":
                    cursor.execute("""
                        UPDATE item_locations
                        SET quantity = quantity + ?, updated_at = datetime('now')
                        WHERE item_name = ? AND location_id = ?
                    """, (adjustment.quantity, adjustment.item_name, adjustment.location_id))
                else:
                    cursor.execute("""
                        UPDATE item_locations
                        SET quantity = quantity - ?, updated_at = datetime('now')
                        WHERE item_name = ? AND location_id = ?
                    """, (adjustment.quantity, adjustment.item_name, adjustment.location_id))

            # Update batch quantity if specified
            if adjustment.batch_id:
                if adjustment.adjustment_type == "increase":
                    cursor.execute("""
                        UPDATE batches
                        SET quantity = quantity + ?, updated_at = datetime('now')
                        WHERE id = ?
                    """, (adjustment.quantity, adjustment.batch_id))
                else:
                    cursor.execute("""
                        UPDATE batches
                        SET quantity = quantity - ?, updated_at = datetime('now')
                        WHERE id = ?
                    """, (adjustment.quantity, adjustment.batch_id))

            return {"message": "Stock adjustment created successfully"}
    except Exception as e:
        logging.error(f"Error creating stock adjustment: {e}")
        raise HTTPException(status_code=500, detail="Error creating stock adjustment")

# ============================================================================
# ALERTS AND NOTIFICATIONS ENDPOINTS
# ============================================================================

@app.get("/alerts")
async def get_alerts(
    unread_only: bool = False,
    alert_type: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get alerts/notifications"""
    try:
        with db_connection.get_cursor() as cursor:
            conditions = []
            params = []

            if unread_only:
                conditions.append("is_read = 0")

            if alert_type:
                conditions.append("alert_type = ?")
                params.append(alert_type)

            where_clause = f"WHERE {' AND '.join(conditions)}" if conditions else ""

            query = f"""
                SELECT * FROM alerts
                {where_clause}
                ORDER BY
                    CASE severity
                        WHEN 'critical' THEN 1
                        WHEN 'high' THEN 2
                        WHEN 'medium' THEN 3
                        WHEN 'low' THEN 4
                    END,
                    created_at DESC
            """

            cursor.execute(query, params)

            alerts = []
            for row in cursor.fetchall():
                alerts.append({
                    "id": row['id'],
                    "alert_type": row['alert_type'],
                    "severity": row['severity'],
                    "item_name": row['item_name'],
                    "location_id": row['location_id'],
                    "batch_id": row['batch_id'],
                    "message": row['message'],
                    "is_read": bool(row['is_read']),
                    "is_resolved": bool(row['is_resolved']),
                    "resolved_by": row['resolved_by'],
                    "resolved_at": row['resolved_at'],
                    "created_at": row['created_at']
                })

            return {"alerts": alerts, "total": len(alerts)}
    except Exception as e:
        logging.error(f"Error fetching alerts: {e}")
        raise HTTPException(status_code=500, detail="Error fetching alerts")

@app.put("/alerts/{alert_id}")
async def update_alert(
    alert_id: int,
    alert_update: AlertUpdate,
    current_user: User = Depends(get_current_user)
):
    """Mark alert as read or resolved"""
    try:
        with db_connection.get_cursor() as cursor:
            updates = []
            params = []

            if alert_update.is_read is not None:
                updates.append("is_read = ?")
                params.append(1 if alert_update.is_read else 0)

            if alert_update.is_resolved is not None:
                updates.append("is_resolved = ?")
                params.append(1 if alert_update.is_resolved else 0)
                if alert_update.is_resolved:
                    updates.append("resolved_at = datetime('now')")
                    if alert_update.resolved_by:
                        updates.append("resolved_by = ?")
                        params.append(alert_update.resolved_by)

            if not updates:
                return {"message": "No updates provided"}

            params.append(alert_id)
            query = f"UPDATE alerts SET {', '.join(updates)} WHERE id = ?"
            cursor.execute(query, params)

            return {"message": "Alert updated successfully"}
    except Exception as e:
        logging.error(f"Error updating alert: {e}")
        raise HTTPException(status_code=500, detail="Error updating alert")

@app.post("/alerts/check-reorder-levels")
async def check_reorder_levels(current_user: User = Depends(get_current_user)):
    """Check all items and create alerts for low stock"""
    try:
        with db_connection.get_cursor() as cursor:
            # Find items below reorder level
            cursor.execute("""
                SELECT item_name, quantity, reorder_level, reorder_quantity
                FROM items
                WHERE quantity <= reorder_level AND reorder_level IS NOT NULL
            """)

            alerts_created = 0
            for row in cursor.fetchall():
                # Check if alert already exists and is unresolved
                cursor.execute("""
                    SELECT id FROM alerts
                    WHERE item_name = ? AND alert_type = 'reorder' AND is_resolved = 0
                """, (row['item_name'],))

                if not cursor.fetchone():
                    severity = 'critical' if row['quantity'] == 0 else 'high' if row['quantity'] < row['reorder_level'] / 2 else 'medium'
                    cursor.execute("""
                        INSERT INTO alerts (alert_type, severity, item_name, message)
                        VALUES ('reorder', ?, ?, ?)
                    """, (
                        severity,
                        row['item_name'],
                        f"Item {row['item_name']} is at {row['quantity']} units (reorder level: {row['reorder_level']})"
                    ))
                    alerts_created += 1

            return {"message": f"Created {alerts_created} reorder alerts"}
    except Exception as e:
        logging.error(f"Error checking reorder levels: {e}")
        raise HTTPException(status_code=500, detail="Error checking reorder levels")

# ============================================================================
# Reports Endpoints
# ============================================================================

@app.get("/reports/low-stock")
async def get_low_stock_report(threshold: int = 10, current_user: User = Depends(get_current_user)):
    """Get low stock report with configurable threshold"""
    try:
        low_stock_items = inventory_service.check_low_stock(threshold)
        return {
            "low_stock_items": low_stock_items,
            "threshold": threshold,
            "count": len(low_stock_items)
        }
    except Exception as e:
        logging.error(f"Error fetching low stock report: {e}")
        raise HTTPException(status_code=500, detail="Error generating report")

@app.get("/reports/inventory")
async def get_inventory_report(groups: Optional[str] = None, current_user: User = Depends(get_current_user)):
    """Get complete inventory report with optional group filtering"""
    group_list = groups.split(',') if groups else None
    report = inventory_service.generate_report(group_list)
    return report

@app.get("/reports/activity")
async def get_activity_report(limit: int = 100, current_user: User = Depends(get_current_user)):
    """Get recent activity"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("""
                SELECT action, item_name, quantity, group_name, timestamp
                FROM history
                ORDER BY timestamp DESC
                LIMIT ?
            """, (limit,))
            activities = []
            for row in cursor.fetchall():
                activities.append({
                    "action": row['action'],
                    "item_name": row['item_name'],
                    "quantity": row['quantity'],
                    "group_name": row['group_name'],
                    "timestamp": row['timestamp']
                })
            return {"activities": activities}
    except Exception as e:
        logging.error(f"Error fetching activity report: {e}")
        raise HTTPException(status_code=500, detail="Error generating report")

# ============================================================================
# Backup and Export Endpoints
# ============================================================================

@app.post("/backup")
async def create_backup(current_user: User = Depends(get_admin_user)):
    """Create database backup (admin only)"""
    try:
        backup_file = inventory_service.backup_data()
        return {"message": "Backup created successfully", "filename": backup_file}
    except Exception as e:
        logging.error(f"Error creating backup: {e}")
        raise HTTPException(status_code=500, detail="Error creating backup")

@app.post("/import/csv")
async def import_inventory_csv(
    file: UploadFile = File(...),
    current_user: User = Depends(get_admin_or_editor)
):
    """Import inventory items from CSV file"""
    try:
        # Read file content
        content = await file.read()
        decoded_content = content.decode('utf-8')
        csv_reader = csv.DictReader(io.StringIO(decoded_content))

        imported = []
        updated = []
        failed = []

        with db_connection.get_cursor() as cursor:
            for row in csv_reader:
                try:
                    item_name = row.get('name') or row.get('item_name')
                    if not item_name:
                        failed.append({"row": row, "reason": "Missing item name"})
                        continue

                    quantity = int(row.get('quantity', 0))
                    group_name = row.get('group') or row.get('group_name')
                    reorder_level = int(row.get('reorder_level', 10)) if row.get('reorder_level') else 10
                    reorder_quantity = int(row.get('reorder_quantity', 50)) if row.get('reorder_quantity') else 50

                    # Check if item exists
                    cursor.execute("SELECT 1 FROM items WHERE item_name = ?", (item_name,))
                    exists = cursor.fetchone()

                    if exists:
                        # Update existing item
                        cursor.execute("""
                            UPDATE items
                            SET quantity = ?, group_name = ?, reorder_level = ?, reorder_quantity = ?
                            WHERE item_name = ?
                        """, (quantity, group_name, reorder_level, reorder_quantity, item_name))
                        updated.append(item_name)
                    else:
                        # Create group if it doesn't exist
                        if group_name:
                            cursor.execute("""
                                INSERT OR IGNORE INTO groups (group_name) VALUES (?)
                            """, (group_name,))

                        # Insert new item
                        cursor.execute("""
                            INSERT INTO items (item_name, quantity, group_name, reorder_level, reorder_quantity)
                            VALUES (?, ?, ?, ?, ?)
                        """, (item_name, quantity, group_name, reorder_level, reorder_quantity))
                        imported.append(item_name)

                    # Add to history
                    cursor.execute("""
                        INSERT INTO history (action, item_name, quantity, group_name, user_name)
                        VALUES (?, ?, ?, ?, ?)
                    """, ("csv_import", item_name, quantity, group_name, current_user.username))

                except Exception as e:
                    failed.append({"row": row, "reason": str(e)})

        return {
            "message": f"Import complete: {len(imported)} new, {len(updated)} updated",
            "imported": imported,
            "updated": updated,
            "failed": failed
        }
    except Exception as e:
        logging.error(f"Error importing CSV: {e}")
        raise HTTPException(status_code=500, detail=f"Error importing CSV: {str(e)}")

@app.post("/import/excel")
async def import_inventory_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_admin_or_editor)
):
    """Import inventory items from Excel file"""
    try:
        from openpyxl import load_workbook
        from io import BytesIO

        # Read file content
        content = await file.read()
        workbook = load_workbook(BytesIO(content))
        sheet = workbook.active

        imported = []
        updated = []
        failed = []

        # Get headers from first row
        headers = [cell.value for cell in sheet[1]]

        with db_connection.get_cursor() as cursor:
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Create dict from row
                    row_dict = dict(zip(headers, row))

                    item_name = row_dict.get('Item Name') or row_dict.get('item_name')
                    if not item_name:
                        failed.append({"row": row_idx, "reason": "Missing item name"})
                        continue

                    quantity = int(row_dict.get('Quantity', 0))
                    group_name = row_dict.get('Group') or row_dict.get('group_name')
                    reorder_level = int(row_dict.get('Reorder Level', 10)) if row_dict.get('Reorder Level') else 10
                    unit = row_dict.get('Unit', 'units')
                    location = row_dict.get('Location', '')
                    description = row_dict.get('Description', '')

                    # Check if item exists
                    cursor.execute("SELECT 1 FROM items WHERE item_name = ?", (item_name,))
                    exists = cursor.fetchone()

                    if exists:
                        # Update existing
                        cursor.execute(
                            """UPDATE items SET quantity = ?, group_name = ?, reorder_point = ?,
                               unit = ?, location = ?, description = ?, updated_at = ?
                               WHERE item_name = ?""",
                            (quantity, group_name, reorder_level, unit, location, description,
                             datetime.now().isoformat(), item_name)
                        )
                        updated.append(item_name)
                    else:
                        # Insert new
                        cursor.execute(
                            """INSERT INTO items
                               (item_name, quantity, group_name, reorder_point, unit, location, description, created_at, updated_at)
                               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                            (item_name, quantity, group_name, reorder_level, unit, location, description,
                             datetime.now().isoformat(), datetime.now().isoformat())
                        )
                        imported.append(item_name)

                        # Log to history
                        cursor.execute(
                            """INSERT INTO history (action, item_name, quantity, group_name, user, timestamp)
                               VALUES (?, ?, ?, ?, ?, ?)""",
                            ('added', item_name, quantity, group_name, current_user.username, datetime.now().isoformat())
                        )

                except Exception as row_error:
                    failed.append({"row": row_idx, "reason": str(row_error)})
                    continue

        logging.info(f"Excel import completed: {len(imported)} imported, {len(updated)} updated, {len(failed)} failed")

        return {
            "status": "success",
            "imported": len(imported),
            "updated": len(updated),
            "failed": len(failed),
            "failed_items": failed[:10] if failed else []  # Return first 10 failures
        }

    except Exception as e:
        logging.error(f"Error importing Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Error importing Excel: {str(e)}")

@app.get("/export/csv")
async def export_inventory_csv(
    groups: Optional[str] = None,
    current_user: User = Depends(get_admin_user)
):
    """Export inventory to CSV file (admin only)"""
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"inventory_export_{timestamp}.csv"
        group_list = groups.split(',') if groups else None

        success = inventory_service.export_to_csv(filename, group_list)
        if success:
            return FileResponse(
                filename,
                media_type='text/csv',
                filename=filename
            )
        raise HTTPException(status_code=500, detail="Failed to export data")
    except Exception as e:
        logging.error(f"Error exporting CSV: {e}")
        raise HTTPException(status_code=500, detail="Error exporting data")

@app.get("/export/excel")
async def export_inventory_excel(
    groups: Optional[str] = None,
    current_user: User = Depends(get_admin_user)
):
    """Export inventory to Excel file (admin only)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"inventory_export_{timestamp}.xlsx"
        group_list = groups.split(',') if groups else None

        # Get inventory data
        conn = db.get_connection()
        cursor = conn.cursor()

        if group_list:
            placeholders = ','.join('?' * len(group_list))
            query = f"SELECT * FROM items WHERE group_name IN ({placeholders})"
            cursor.execute(query, group_list)
        else:
            cursor.execute("SELECT * FROM items")

        items = cursor.fetchall()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"

        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Add headers
        headers = ["Item Name", "Quantity", "Group", "Reorder Level", "Unit", "Location", "Description", "Created At"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Add data
        for row_num, item in enumerate(items, 2):
            ws.cell(row=row_num, column=1, value=item[0])  # item_name
            ws.cell(row=row_num, column=2, value=item[1])  # quantity
            ws.cell(row=row_num, column=3, value=item[2])  # group_name
            ws.cell(row=row_num, column=4, value=item[3])  # reorder_point
            ws.cell(row=row_num, column=5, value=item[4])  # unit
            ws.cell(row=row_num, column=6, value=item[5])  # location
            ws.cell(row=row_num, column=7, value=item[6])  # description
            ws.cell(row=row_num, column=8, value=item[7])  # created_at

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save file
        wb.save(filename)
        conn.close()

        return FileResponse(
            filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=filename
        )
    except Exception as e:
        logging.error(f"Error exporting Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Error exporting data: {str(e)}")

@app.get("/export/pdf")
async def export_inventory_pdf(
    groups: Optional[str] = None,
    current_user: User = Depends(get_admin_user)
):
    """Export inventory to PDF file (admin only)"""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"inventory_report_{timestamp}.pdf"
        group_list = groups.split(',') if groups else None

        # Get inventory data
        conn = db.get_connection()
        cursor = conn.cursor()

        if group_list:
            placeholders = ','.join('?' * len(group_list))
            query = f"SELECT item_name, quantity, group_name, reorder_point, unit FROM items WHERE group_name IN ({placeholders})"
            cursor.execute(query, group_list)
        else:
            cursor.execute("SELECT item_name, quantity, group_name, reorder_point, unit FROM items")

        items = cursor.fetchall()
        conn.close()

        # Create PDF
        doc = SimpleDocTemplate(filename, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1976d2'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        title = Paragraph("Inventory Management Report", title_style)
        elements.append(title)

        # Subtitle with date
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.grey,
            spaceAfter=20,
            alignment=TA_CENTER
        )
        subtitle = Paragraph(f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", subtitle_style)
        elements.append(subtitle)
        elements.append(Spacer(1, 0.3*inch))

        # Summary
        summary_text = f"Total Items: <b>{len(items)}</b>"
        summary = Paragraph(summary_text, styles['Normal'])
        elements.append(summary)
        elements.append(Spacer(1, 0.2*inch))

        # Table data
        data = [['Item Name', 'Quantity', 'Group', 'Reorder Level', 'Unit']]
        for item in items:
            data.append([
                item[0] or '',
                str(item[1]) if item[1] is not None else '0',
                item[2] or 'N/A',
                str(item[3]) if item[3] is not None else 'N/A',
                item[4] or 'N/A'
            ])

        # Create table
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ]))

        elements.append(table)
        doc.build(elements)

        return FileResponse(
            filename,
            media_type='application/pdf',
            filename=filename
        )
    except Exception as e:
        logging.error(f"Error exporting PDF: {e}")
        raise HTTPException(status_code=500, detail=f"Error exporting PDF: {str(e)}")

# ============================================================================
# QR CODE GENERATION ENDPOINTS
# ============================================================================

@app.get("/items/{item_name}/qrcode")
async def generate_item_qrcode(
    item_name: str,
    size: int = 10,
    current_user: User = Depends(get_current_user)
):
    """Generate QR code for an item"""
    try:
        import qrcode
        from io import BytesIO

        # Get item details
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM items WHERE item_name = ?", (item_name,))
        item = cursor.fetchone()
        conn.close()

        if not item:
            raise HTTPException(status_code=404, detail="Item not found")

        # Create QR code data
        qr_data = {
            "item_name": item[0],
            "quantity": item[1],
            "group": item[2],
            "reorder_point": item[3],
        }
        import json
        qr_content = json.dumps(qr_data)

        # Generate QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=size,
            border=4,
        )
        qr.add_data(qr_content)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")

        # Save to bytes
        img_byte_arr = BytesIO()
        img.save(img_byte_arr, format='PNG')
        img_byte_arr.seek(0)

        return Response(content=img_byte_arr.getvalue(), media_type="image/png")
    except Exception as e:
        logging.error(f"Error generating QR code: {e}")
        raise HTTPException(status_code=500, detail=f"Error generating QR code: {str(e)}")

@app.get("/items/{item_name}/barcode")
async def generate_item_barcode(
    item_name: str,
    current_user: User = Depends(get_current_user)
):
    """Generate barcode for an item (Code128)"""
    try:
        from barcode import Code128
        from barcode.writer import ImageWriter
        from io import BytesIO

        # Create barcode
        # Use item name or a unique identifier
        barcode_data = item_name.replace(" ", "_")[:20]  # Limit length

        barcode = Code128(barcode_data, writer=ImageWriter())

        # Save to bytes
        img_byte_arr = BytesIO()
        barcode.write(img_byte_arr)
        img_byte_arr.seek(0)

        return Response(content=img_byte_arr.getvalue(), media_type="image/png")
    except ImportError:
        raise HTTPException(
            status_code=501,
            detail="Barcode generation not available. Install python-barcode library."
        )
    except Exception as e:
        logging.error(f"Error generating barcode: {e}")
        raise HTTPException(status_code=500, detail=f"Error generating barcode: {str(e)}")

@app.post("/items/{item_name}/print-label")
async def generate_print_label(
    item_name: str,
    include_qrcode: bool = True,
    current_user: User = Depends(get_current_user)
):
    """Generate printable label with QR code"""
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Image
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        import qrcode
        from io import BytesIO

        # Get item details
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM items WHERE item_name = ?", (item_name,))
        item = cursor.fetchone()
        conn.close()

        if not item:
            raise HTTPException(status_code=404, detail="Item not found")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"label_{item_name}_{timestamp}.pdf"

        # Create PDF
        doc = SimpleDocTemplate(
            filename,
            pagesize=(4*inch, 2*inch),  # Label size
            topMargin=0.25*inch,
            bottomMargin=0.25*inch,
            leftMargin=0.25*inch,
            rightMargin=0.25*inch
        )
        elements = []
        styles = getSampleStyleSheet()

        if include_qrcode:
            # Generate QR code
            import json
            qr_data = json.dumps({
                "item_name": item[0],
                "quantity": item[1],
                "group": item[2]
            })
            qr = qrcode.QRCode(version=1, box_size=10, border=2)
            qr.add_data(qr_data)
            qr.make(fit=True)
            qr_img = qr.make_image(fill_color="black", back_color="white")

            # Save QR to bytes
            qr_byte_arr = BytesIO()
            qr_img.save(qr_byte_arr, format='PNG')
            qr_byte_arr.seek(0)

            # Create layout
            data = [
                [
                    Image(qr_byte_arr, width=1.2*inch, height=1.2*inch),
                    [
                        Paragraph(f"<b>{item[0]}</b>", styles['Heading2']),
                        Paragraph(f"Qty: {item[1]}", styles['Normal']),
                        Paragraph(f"Group: {item[2] or 'N/A'}", styles['Normal']),
                        Paragraph(f"Reorder: {item[3] or 'N/A'}", styles['Normal']),
                    ]
                ]
            ]
        else:
            data = [
                [
                    Paragraph(f"<b>{item[0]}</b>", styles['Heading1']),
                ],
                [
                    Paragraph(f"Quantity: {item[1]}", styles['Normal']),
                ],
                [
                    Paragraph(f"Group: {item[2] or 'N/A'}", styles['Normal']),
                ]
            ]

        table = Table(data)
        table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ]))

        elements.append(table)
        doc.build(elements)

        return FileResponse(filename, media_type='application/pdf', filename=filename)
    except Exception as e:
        logging.error(f"Error generating label: {e}")
        raise HTTPException(status_code=500, detail=f"Error generating label: {str(e)}")

# ============================================================================
# PURCHASE ORDER ENDPOINTS
# ============================================================================

class PurchaseOrderItem(BaseModel):
    item_name: str
    quantity: int
    unit_price: float
    notes: Optional[str] = None

class PurchaseOrder(BaseModel):
    supplier_id: int
    location_id: Optional[int] = None
    order_date: Optional[str] = None
    expected_delivery_date: Optional[str] = None
    status: str = 'pending'  # pending, confirmed, shipped, received, cancelled
    items: List[PurchaseOrderItem]
    notes: Optional[str] = None

@app.post("/purchase-orders")
async def create_purchase_order(
    po: PurchaseOrder,
    current_user: User = Depends(get_admin_or_editor)
):
    """Create a new purchase order"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # Validate supplier
        cursor.execute("SELECT name FROM suppliers WHERE id = ?", (po.supplier_id,))
        supplier = cursor.fetchone()
        if not supplier:
            raise HTTPException(status_code=404, detail="Supplier not found")

        # Calculate total
        total_amount = sum(item.quantity * item.unit_price for item in po.items)

        # Create purchase order
        order_number = f"PO-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        order_date = po.order_date or datetime.now().isoformat()

        cursor.execute(
            """INSERT INTO purchase_orders
               (order_number, supplier_id, location_id, order_date, expected_delivery_date,
                status, total_amount, created_by, created_at, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (order_number, po.supplier_id, po.location_id, order_date,
             po.expected_delivery_date, po.status, total_amount,
             current_user.username, datetime.now().isoformat(), po.notes)
        )

        po_id = cursor.lastrowid

        # Add line items
        for item in po.items:
            cursor.execute(
                """INSERT INTO purchase_order_items
                   (po_id, item_name, quantity, unit_price, total_price, notes)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (po_id, item.item_name, item.quantity, item.unit_price,
                 item.quantity * item.unit_price, item.notes)
            )

        # Log to history
        cursor.execute(
            """INSERT INTO history (action, item_name, user, timestamp, notes)
               VALUES (?, ?, ?, ?, ?)""",
            ('purchase_order_created', 'Multiple Items', current_user.username,
             datetime.now().isoformat(), f"PO: {order_number}, Supplier: {supplier[0]}")
        )

        conn.commit()
        conn.close()

        logging.info(f"Purchase order created: {order_number}")

        return {
            "status": "success",
            "order_number": order_number,
            "po_id": po_id,
            "total_amount": total_amount
        }

    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error creating purchase order: {e}")
        raise HTTPException(status_code=500, detail=f"Error creating purchase order: {str(e)}")

@app.get("/purchase-orders")
async def get_purchase_orders(
    status: Optional[str] = None,
    supplier_id: Optional[int] = None,
    limit: int = 50,
    current_user: User = Depends(get_current_user)
):
    """Get purchase orders"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        query = """
            SELECT po.*, s.name as supplier_name, l.name as location_name
            FROM purchase_orders po
            LEFT JOIN suppliers s ON po.supplier_id = s.id
            LEFT JOIN locations l ON po.location_id = l.id
            WHERE 1=1
        """
        params = []

        if status:
            query += " AND po.status = ?"
            params.append(status)

        if supplier_id:
            query += " AND po.supplier_id = ?"
            params.append(supplier_id)

        query += " ORDER BY po.created_at DESC LIMIT ?"
        params.append(limit)

        cursor.execute(query, params)
        orders = cursor.fetchall()

        # Get items for each order
        result = []
        for order in orders:
            cursor.execute(
                """SELECT * FROM purchase_order_items WHERE po_id = ?""",
                (order[0],)
            )
            items = cursor.fetchall()

            result.append({
                "id": order[0],
                "order_number": order[1],
                "supplier_id": order[2],
                "supplier_name": order[-2],
                "location_id": order[3],
                "location_name": order[-1],
                "order_date": order[4],
                "expected_delivery_date": order[5],
                "actual_delivery_date": order[6],
                "status": order[7],
                "total_amount": order[8],
                "created_by": order[9],
                "created_at": order[10],
                "updated_at": order[11],
                "notes": order[12],
                "items": [
                    {
                        "id": item[0],
                        "item_name": item[2],
                        "quantity": item[3],
                        "unit_price": item[4],
                        "total_price": item[5],
                        "received_quantity": item[6],
                        "notes": item[7]
                    }
                    for item in items
                ]
            })

        conn.close()
        return {"purchase_orders": result}

    except Exception as e:
        logging.error(f"Error fetching purchase orders: {e}")
        raise HTTPException(status_code=500, detail=f"Error fetching purchase orders: {str(e)}")

@app.put("/purchase-orders/{po_id}/status")
async def update_purchase_order_status(
    po_id: int,
    status: str,
    current_user: User = Depends(get_admin_or_editor)
):
    """Update purchase order status"""
    try:
        valid_statuses = ['pending', 'confirmed', 'shipped', 'received', 'cancelled']
        if status not in valid_statuses:
            raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")

        conn = db.get_connection()
        cursor = conn.cursor()

        cursor.execute(
            """UPDATE purchase_orders
               SET status = ?, updated_at = ?
               WHERE id = ?""",
            (status, datetime.now().isoformat(), po_id)
        )

        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Purchase order not found")

        conn.commit()
        conn.close()

        return {"status": "success", "new_status": status}

    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error updating purchase order status: {e}")
        raise HTTPException(status_code=500, detail=f"Error updating status: {str(e)}")

@app.post("/purchase-orders/{po_id}/receive")
async def receive_purchase_order(
    po_id: int,
    received_items: List[dict],  # [{"item_name": "...", "quantity": int}]
    current_user: User = Depends(get_admin_or_editor)
):
    """Receive items from a purchase order and update inventory"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # Get PO details
        cursor.execute("SELECT * FROM purchase_orders WHERE id = ?", (po_id,))
        po = cursor.fetchone()
        if not po:
            raise HTTPException(status_code=404, detail="Purchase order not found")

        # Update inventory for received items
        for item in received_items:
            item_name = item['item_name']
            quantity = item['quantity']

            # Update main inventory
            cursor.execute(
                """UPDATE items SET quantity = quantity + ?, updated_at = ?
                   WHERE item_name = ?""",
                (quantity, datetime.now().isoformat(), item_name)
            )

            # Update received quantity in PO items
            cursor.execute(
                """UPDATE purchase_order_items
                   SET received_quantity = COALESCE(received_quantity, 0) + ?
                   WHERE po_id = ? AND item_name = ?""",
                (quantity, po_id, item_name)
            )

            # Log to history
            cursor.execute(
                """INSERT INTO history (action, item_name, quantity, user, timestamp, notes)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                ('received_from_po', item_name, quantity, current_user.username,
                 datetime.now().isoformat(), f"PO ID: {po_id}")
            )

        # Update PO status
        cursor.execute(
            """UPDATE purchase_orders
               SET status = 'received', actual_delivery_date = ?, updated_at = ?
               WHERE id = ?""",
            (datetime.now().isoformat(), datetime.now().isoformat(), po_id)
        )

        conn.commit()
        conn.close()

        return {"status": "success", "message": "Items received and inventory updated"}

    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error receiving purchase order: {e}")
        raise HTTPException(status_code=500, detail=f"Error receiving PO: {str(e)}")

# ============================================================================
# STOCK TRANSFER ENDPOINTS
# ============================================================================

class StockTransfer(BaseModel):
    item_name: str
    from_location_id: int
    to_location_id: int
    quantity: int
    transfer_date: Optional[str] = None
    notes: Optional[str] = None
    transferred_by: Optional[str] = None

@app.post("/stock-transfers")
async def create_stock_transfer(
    transfer: StockTransfer,
    current_user: User = Depends(get_admin_or_editor)
):
    """Transfer stock between locations"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # Validate locations
        cursor.execute("SELECT id, name FROM locations WHERE id IN (?, ?)",
                      (transfer.from_location_id, transfer.to_location_id))
        locations = cursor.fetchall()

        if len(locations) != 2:
            raise HTTPException(status_code=404, detail="One or both locations not found")

        # Check stock at source location
        cursor.execute(
            """SELECT quantity FROM item_locations
               WHERE item_name = ? AND location_id = ?""",
            (transfer.item_name, transfer.from_location_id)
        )
        source_stock = cursor.fetchone()

        if not source_stock or source_stock[0] < transfer.quantity:
            raise HTTPException(
                status_code=400,
                detail=f"Insufficient stock at source location. Available: {source_stock[0] if source_stock else 0}"
            )

        # Update source location
        new_source_qty = source_stock[0] - transfer.quantity
        cursor.execute(
            """UPDATE item_locations SET quantity = ?, updated_at = ?
               WHERE item_name = ? AND location_id = ?""",
            (new_source_qty, datetime.now().isoformat(), transfer.item_name, transfer.from_location_id)
        )

        # Update or insert destination location
        cursor.execute(
            """SELECT quantity FROM item_locations
               WHERE item_name = ? AND location_id = ?""",
            (transfer.item_name, transfer.to_location_id)
        )
        dest_stock = cursor.fetchone()

        if dest_stock:
            new_dest_qty = dest_stock[0] + transfer.quantity
            cursor.execute(
                """UPDATE item_locations SET quantity = ?, updated_at = ?
                   WHERE item_name = ? AND location_id = ?""",
                (new_dest_qty, datetime.now().isoformat(), transfer.item_name, transfer.to_location_id)
            )
        else:
            cursor.execute(
                """INSERT INTO item_locations (item_name, location_id, quantity, created_at, updated_at)
                   VALUES (?, ?, ?, ?, ?)""",
                (transfer.item_name, transfer.to_location_id, transfer.quantity,
                 datetime.now().isoformat(), datetime.now().isoformat())
            )

        # Log the transfer in stock adjustments
        cursor.execute(
            """INSERT INTO stock_adjustments
               (item_name, adjustment_type, quantity, reason, location_id, reference_number, notes, adjusted_by, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (transfer.item_name, 'subtraction', transfer.quantity, 'transfer',
             transfer.from_location_id, f"Transfer to Location {transfer.to_location_id}",
             transfer.notes, current_user.username, datetime.now().isoformat())
        )

        cursor.execute(
            """INSERT INTO stock_adjustments
               (item_name, adjustment_type, quantity, reason, location_id, reference_number, notes, adjusted_by, created_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (transfer.item_name, 'addition', transfer.quantity, 'transfer',
             transfer.to_location_id, f"Transfer from Location {transfer.from_location_id}",
             transfer.notes, current_user.username, datetime.now().isoformat())
        )

        # Log to history
        cursor.execute(
            """INSERT INTO history (action, item_name, quantity, user, timestamp, notes)
               VALUES (?, ?, ?, ?, ?, ?)""",
            ('transfer', transfer.item_name, transfer.quantity, current_user.username,
             datetime.now().isoformat(),
             f"From Location {transfer.from_location_id} to {transfer.to_location_id}")
        )

        conn.commit()
        conn.close()

        logging.info(f"Stock transfer completed: {transfer.item_name}, {transfer.quantity} units from {transfer.from_location_id} to {transfer.to_location_id}")

        return {
            "status": "success",
            "message": "Stock transferred successfully",
            "transfer": transfer.dict()
        }

    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error transferring stock: {e}")
        raise HTTPException(status_code=500, detail=f"Error transferring stock: {str(e)}")

@app.get("/stock-transfers")
async def get_stock_transfers(
    item_name: Optional[str] = None,
    location_id: Optional[int] = None,
    limit: int = 50,
    current_user: User = Depends(get_current_user)
):
    """Get stock transfer history"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        query = """
            SELECT sa.*, l.name as location_name
            FROM stock_adjustments sa
            LEFT JOIN locations l ON sa.location_id = l.id
            WHERE sa.reason = 'transfer'
        """
        params = []

        if item_name:
            query += " AND sa.item_name = ?"
            params.append(item_name)

        if location_id:
            query += " AND sa.location_id = ?"
            params.append(location_id)

        query += " ORDER BY sa.created_at DESC LIMIT ?"
        params.append(limit)

        cursor.execute(query, params)
        transfers = cursor.fetchall()
        conn.close()

        return {
            "transfers": [
                {
                    "id": t[0],
                    "item_name": t[1],
                    "adjustment_type": t[2],
                    "quantity": t[3],
                    "location_id": t[5],
                    "location_name": t[-1],
                    "reference_number": t[7],
                    "notes": t[8],
                    "adjusted_by": t[9],
                    "created_at": t[11]
                }
                for t in transfers
            ]
        }

    except Exception as e:
        logging.error(f"Error fetching stock transfers: {e}")
        raise HTTPException(status_code=500, detail=f"Error fetching stock transfers: {str(e)}")

# ============================================================================
# NOTES/COMMENTS ENDPOINTS
# ============================================================================

class NoteCreate(BaseModel):
    item_name: str
    note_text: str
    is_pinned: Optional[bool] = False

class NoteUpdate(BaseModel):
    note_text: Optional[str] = None
    is_pinned: Optional[bool] = None

@app.get("/notes/{item_name}")
async def get_item_notes(item_name: str, current_user: User = Depends(get_current_user)):
    """Get all notes for a specific item"""
    try:
        with db_connection.get_cursor() as cursor:
            cursor.execute("""
                SELECT id, item_name, note_text, created_by, created_at, updated_at, is_pinned
                FROM notes
                WHERE item_name = ?
                ORDER BY is_pinned DESC, created_at DESC
            """, (item_name,))
            notes = cursor.fetchall()
            return notes
    except Exception as e:
        logging.error(f"Error getting notes: {e}")
        raise HTTPException(status_code=500, detail="Error retrieving notes")

@app.post("/notes")
async def create_note(note: NoteCreate, current_user: User = Depends(get_current_user)):
    """Create a new note for an item"""
    if current_user.role == 'viewer':
        raise HTTPException(status_code=403, detail="Viewers cannot create notes")

    try:
        with db_connection.get_cursor() as cursor:
            # Check if item exists
            cursor.execute("SELECT 1 FROM items WHERE item_name = ?", (note.item_name,))
            if not cursor.fetchone():
                raise HTTPException(status_code=404, detail="Item not found")

            cursor.execute("""
                INSERT INTO notes (item_name, note_text, created_by, is_pinned)
                VALUES (?, ?, ?, ?)
            """, (note.item_name, note.note_text, current_user.username, 1 if note.is_pinned else 0))

            note_id = cursor.lastrowid

            # Add to history
            cursor.execute("""
                INSERT INTO history (action, item_name, user_name)
                VALUES (?, ?, ?)
            """, ("note_added", note.item_name, current_user.username))

            return {"id": note_id, "message": "Note created successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error creating note: {e}")
        raise HTTPException(status_code=500, detail="Error creating note")

@app.put("/notes/{note_id}")
async def update_note(note_id: int, note: NoteUpdate, current_user: User = Depends(get_current_user)):
    """Update an existing note"""
    if current_user.role == 'viewer':
        raise HTTPException(status_code=403, detail="Viewers cannot update notes")

    try:
        with db_connection.get_cursor() as cursor:
            # Check if note exists and get creator
            cursor.execute("SELECT created_by, item_name FROM notes WHERE id = ?", (note_id,))
            existing = cursor.fetchone()
            if not existing:
                raise HTTPException(status_code=404, detail="Note not found")

            # Only admin or creator can edit
            if current_user.role != 'admin' and existing['created_by'] != current_user.username:
                raise HTTPException(status_code=403, detail="Cannot edit others' notes")

            updates = []
            params = []
            if note.note_text is not None:
                updates.append("note_text = ?")
                params.append(note.note_text)
            if note.is_pinned is not None:
                updates.append("is_pinned = ?")
                params.append(1 if note.is_pinned else 0)

            if updates:
                updates.append("updated_at = datetime('now')")
                params.append(note_id)

                cursor.execute(f"""
                    UPDATE notes
                    SET {', '.join(updates)}
                    WHERE id = ?
                """, params)

                return {"message": "Note updated successfully"}
            return {"message": "No changes made"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error updating note: {e}")
        raise HTTPException(status_code=500, detail="Error updating note")

@app.delete("/notes/{note_id}")
async def delete_note(note_id: int, current_user: User = Depends(get_current_user)):
    """Delete a note"""
    if current_user.role == 'viewer':
        raise HTTPException(status_code=403, detail="Viewers cannot delete notes")

    try:
        with db_connection.get_cursor() as cursor:
            # Check if note exists and get creator
            cursor.execute("SELECT created_by FROM notes WHERE id = ?", (note_id,))
            existing = cursor.fetchone()
            if not existing:
                raise HTTPException(status_code=404, detail="Note not found")

            # Only admin or creator can delete
            if current_user.role != 'admin' and existing['created_by'] != current_user.username:
                raise HTTPException(status_code=403, detail="Cannot delete others' notes")

            cursor.execute("DELETE FROM notes WHERE id = ?", (note_id,))
            return {"message": "Note deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error deleting note: {e}")
        raise HTTPException(status_code=500, detail="Error deleting note")

# ============================================================================
# FINANCIAL REPORTING AND ANALYTICS
# ============================================================================

@app.get("/analytics/financial-summary")
async def get_financial_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Get financial summary with revenue, costs, and profit"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # Build date filter
        date_filter = ""
        params = []
        if start_date:
            date_filter += " AND po.order_date >= ?"
            params.append(start_date)
        if end_date:
            date_filter += " AND po.order_date <= ?"
            params.append(end_date)

        # Get total purchase costs from purchase orders
        cursor.execute(
            f"""SELECT
                   COALESCE(SUM(total_amount), 0) as total_purchase_cost,
                   COUNT(DISTINCT id) as total_orders
               FROM purchase_orders po
               WHERE status = 'received'{date_filter}""",
            params
        )
        purchase_data = cursor.fetchone()
        total_purchase_cost = purchase_data['total_purchase_cost'] if purchase_data else 0
        total_orders = purchase_data['total_orders'] if purchase_data else 0

        # Calculate current inventory value based on latest prices
        cursor.execute("""
            SELECT
                COALESCE(SUM(i.quantity * COALESCE(p.price, 0)), 0) as inventory_value,
                COUNT(DISTINCT i.item_name) as total_items
            FROM items i
            LEFT JOIN (
                SELECT item_name, AVG(price) as price
                FROM prices
                GROUP BY item_name
            ) p ON i.item_name = p.item_name
        """)
        inventory_data = cursor.fetchone()
        inventory_value = inventory_data['inventory_value'] if inventory_data else 0
        total_items = inventory_data['total_items'] if inventory_data else 0

        # Get stock adjustments for revenue estimation (returned items)
        cursor.execute(
            f"""SELECT
                   COUNT(*) as total_adjustments,
                   COALESCE(SUM(CASE WHEN adjustment_type = 'increase' THEN quantity ELSE 0 END), 0) as items_added,
                   COALESCE(SUM(CASE WHEN adjustment_type = 'decrease' THEN quantity ELSE 0 END), 0) as items_removed
               FROM stock_adjustments
               WHERE 1=1{date_filter.replace('po.', '')}""",
            params
        )
        adjustments = cursor.fetchone()

        # Calculate estimated revenue (items sold * average price)
        # Assuming items_removed represents sales/usage
        items_sold = adjustments['items_removed'] if adjustments else 0
        cursor.execute("""
            SELECT AVG(price) as avg_price
            FROM prices
        """)
        avg_price_data = cursor.fetchone()
        avg_price = avg_price_data['avg_price'] if avg_price_data and avg_price_data['avg_price'] else 0
        estimated_revenue = items_sold * avg_price

        # Calculate profit margin
        gross_profit = estimated_revenue - total_purchase_cost
        profit_margin = (gross_profit / estimated_revenue * 100) if estimated_revenue > 0 else 0

        conn.close()

        return {
            "total_purchase_cost": round(total_purchase_cost, 2),
            "estimated_revenue": round(estimated_revenue, 2),
            "gross_profit": round(gross_profit, 2),
            "profit_margin_percent": round(profit_margin, 2),
            "current_inventory_value": round(inventory_value, 2),
            "total_items": total_items,
            "total_purchase_orders": total_orders,
            "items_sold": items_sold,
            "items_added": adjustments['items_added'] if adjustments else 0,
            "total_adjustments": adjustments['total_adjustments'] if adjustments else 0
        }

    except Exception as e:
        logging.error(f"Error getting financial summary: {e}")
        raise HTTPException(status_code=500, detail=f"Error getting financial summary: {str(e)}")


@app.get("/analytics/inventory-value")
async def get_inventory_value_breakdown(
    current_user: User = Depends(get_current_user)
):
    """Get inventory value broken down by group/category"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                COALESCE(i.group_name, 'Uncategorized') as category,
                COUNT(i.item_name) as item_count,
                SUM(i.quantity) as total_quantity,
                COALESCE(SUM(i.quantity * p.avg_price), 0) as total_value,
                COALESCE(AVG(p.avg_price), 0) as avg_unit_price
            FROM items i
            LEFT JOIN (
                SELECT item_name, AVG(price) as avg_price
                FROM prices
                GROUP BY item_name
            ) p ON i.item_name = p.item_name
            GROUP BY i.group_name
            ORDER BY total_value DESC
        """)

        breakdown = []
        for row in cursor.fetchall():
            breakdown.append({
                "category": row['category'],
                "item_count": row['item_count'],
                "total_quantity": row['total_quantity'],
                "total_value": round(row['total_value'], 2),
                "avg_unit_price": round(row['avg_unit_price'], 2)
            })

        conn.close()
        return breakdown

    except Exception as e:
        logging.error(f"Error getting inventory value breakdown: {e}")
        raise HTTPException(status_code=500, detail=f"Error getting inventory value breakdown: {str(e)}")


@app.get("/analytics/top-items")
async def get_top_items(
    metric: str = 'value',  # 'value', 'quantity', 'movement'
    limit: int = 10,
    current_user: User = Depends(get_current_user)
):
    """Get top items by various metrics"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        if metric == 'value':
            # Top items by total value
            cursor.execute("""
                SELECT
                    i.item_name,
                    i.quantity,
                    i.group_name,
                    COALESCE(p.avg_price, 0) as unit_price,
                    i.quantity * COALESCE(p.avg_price, 0) as total_value
                FROM items i
                LEFT JOIN (
                    SELECT item_name, AVG(price) as avg_price
                    FROM prices
                    GROUP BY item_name
                ) p ON i.item_name = p.item_name
                ORDER BY total_value DESC
                LIMIT ?
            """, (limit,))
        elif metric == 'quantity':
            # Top items by quantity in stock
            cursor.execute("""
                SELECT
                    i.item_name,
                    i.quantity,
                    i.group_name,
                    COALESCE(p.avg_price, 0) as unit_price,
                    i.quantity * COALESCE(p.avg_price, 0) as total_value
                FROM items i
                LEFT JOIN (
                    SELECT item_name, AVG(price) as avg_price
                    FROM prices
                    GROUP BY item_name
                ) p ON i.item_name = p.item_name
                ORDER BY i.quantity DESC
                LIMIT ?
            """, (limit,))
        else:  # movement
            # Top items by stock movement (adjustments)
            cursor.execute("""
                SELECT
                    sa.item_name,
                    i.quantity as current_quantity,
                    i.group_name,
                    COUNT(*) as movement_count,
                    SUM(sa.quantity) as total_moved,
                    COALESCE(p.avg_price, 0) as unit_price
                FROM stock_adjustments sa
                JOIN items i ON sa.item_name = i.item_name
                LEFT JOIN (
                    SELECT item_name, AVG(price) as avg_price
                    FROM prices
                    GROUP BY item_name
                ) p ON i.item_name = p.item_name
                WHERE sa.adjustment_date >= datetime('now', '-30 days')
                GROUP BY sa.item_name, i.quantity, i.group_name
                ORDER BY total_moved DESC
                LIMIT ?
            """, (limit,))

        items = []
        for row in cursor.fetchall():
            item_data = {
                "item_name": row['item_name'],
                "group_name": row['group_name'] or 'Uncategorized'
            }

            if metric == 'movement':
                item_data.update({
                    "current_quantity": row['current_quantity'],
                    "movement_count": row['movement_count'],
                    "total_moved": row['total_moved'],
                    "unit_price": round(row['unit_price'], 2)
                })
            else:
                item_data.update({
                    "quantity": row['quantity'],
                    "unit_price": round(row['unit_price'], 2),
                    "total_value": round(row['total_value'], 2)
                })

            items.append(item_data)

        conn.close()
        return items

    except Exception as e:
        logging.error(f"Error getting top items: {e}")
        raise HTTPException(status_code=500, detail=f"Error getting top items: {str(e)}")


@app.get("/analytics/revenue-by-period")
async def get_revenue_by_period(
    period: str = 'daily',  # 'daily', 'weekly', 'monthly'
    limit: int = 30,
    current_user: User = Depends(get_current_user)
):
    """Get revenue/cost trends over time"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # Determine date grouping
        if period == 'daily':
            date_format = '%Y-%m-%d'
            date_trunc = "date(order_date)"
        elif period == 'weekly':
            date_format = '%Y-W%W'
            date_trunc = "strftime('%Y-W%W', order_date)"
        else:  # monthly
            date_format = '%Y-%m'
            date_trunc = "strftime('%Y-%m', order_date)"

        # Get purchase costs by period
        cursor.execute(f"""
            SELECT
                {date_trunc} as period,
                SUM(total_amount) as total_cost,
                COUNT(*) as order_count
            FROM purchase_orders
            WHERE status = 'received'
                AND order_date >= datetime('now', '-{limit} days')
            GROUP BY period
            ORDER BY period DESC
            LIMIT ?
        """, (limit,))

        periods = []
        for row in cursor.fetchall():
            periods.append({
                "period": row['period'],
                "total_cost": round(row['total_cost'], 2) if row['total_cost'] else 0,
                "order_count": row['order_count']
            })

        conn.close()
        return periods

    except Exception as e:
        logging.error(f"Error getting revenue by period: {e}")
        raise HTTPException(status_code=500, detail=f"Error getting revenue by period: {str(e)}")


@app.get("/analytics/cost-analysis")
async def get_cost_analysis(
    current_user: User = Depends(get_current_user)
):
    """Get cost analysis by supplier and item"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # Cost by supplier
        cursor.execute("""
            SELECT
                s.name as supplier_name,
                s.id as supplier_id,
                COUNT(DISTINCT po.id) as total_orders,
                COALESCE(SUM(po.total_amount), 0) as total_spent,
                COALESCE(AVG(po.total_amount), 0) as avg_order_value
            FROM suppliers s
            LEFT JOIN purchase_orders po ON s.id = po.supplier_id AND po.status = 'received'
            GROUP BY s.id, s.name
            HAVING total_orders > 0
            ORDER BY total_spent DESC
        """)

        suppliers = []
        for row in cursor.fetchall():
            suppliers.append({
                "supplier_name": row['supplier_name'],
                "supplier_id": row['supplier_id'],
                "total_orders": row['total_orders'],
                "total_spent": round(row['total_spent'], 2),
                "avg_order_value": round(row['avg_order_value'], 2)
            })

        # Cost by item (from PO items)
        cursor.execute("""
            SELECT
                poi.item_name,
                COUNT(DISTINCT poi.po_id) as order_count,
                SUM(poi.quantity) as total_quantity_ordered,
                COALESCE(AVG(poi.unit_price), 0) as avg_unit_cost,
                COALESCE(SUM(poi.total_price), 0) as total_cost
            FROM purchase_order_items poi
            JOIN purchase_orders po ON poi.po_id = po.id
            WHERE po.status = 'received'
            GROUP BY poi.item_name
            ORDER BY total_cost DESC
            LIMIT 20
        """)

        items = []
        for row in cursor.fetchall():
            items.append({
                "item_name": row['item_name'],
                "order_count": row['order_count'],
                "total_quantity_ordered": row['total_quantity_ordered'],
                "avg_unit_cost": round(row['avg_unit_cost'], 2),
                "total_cost": round(row['total_cost'], 2)
            })

        conn.close()
        return {
            "by_supplier": suppliers,
            "by_item": items
        }

    except Exception as e:
        logging.error(f"Error getting cost analysis: {e}")
        raise HTTPException(status_code=500, detail=f"Error getting cost analysis: {str(e)}")


@app.get("/analytics/profit-margins")
async def get_profit_margins(
    current_user: User = Depends(get_current_user)
):
    """Calculate profit margins by item and category"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # Profit margins by item (comparing purchase cost vs selling price)
        cursor.execute("""
            SELECT
                i.item_name,
                i.group_name,
                COALESCE(AVG(poi.unit_price), 0) as avg_cost,
                COALESCE(AVG(p.price), 0) as avg_selling_price,
                (COALESCE(AVG(p.price), 0) - COALESCE(AVG(poi.unit_price), 0)) as profit_per_unit,
                CASE
                    WHEN COALESCE(AVG(p.price), 0) > 0 THEN
                        ((COALESCE(AVG(p.price), 0) - COALESCE(AVG(poi.unit_price), 0)) / COALESCE(AVG(p.price), 0) * 100)
                    ELSE 0
                END as profit_margin_percent
            FROM items i
            LEFT JOIN purchase_order_items poi ON i.item_name = poi.item_name
            LEFT JOIN prices p ON i.item_name = p.item_name
            GROUP BY i.item_name, i.group_name
            HAVING avg_selling_price > 0 OR avg_cost > 0
            ORDER BY profit_margin_percent DESC
            LIMIT 20
        """)

        items = []
        for row in cursor.fetchall():
            items.append({
                "item_name": row['item_name'],
                "group_name": row['group_name'] or 'Uncategorized',
                "avg_cost": round(row['avg_cost'], 2),
                "avg_selling_price": round(row['avg_selling_price'], 2),
                "profit_per_unit": round(row['profit_per_unit'], 2),
                "profit_margin_percent": round(row['profit_margin_percent'], 2)
            })

        # Profit margins by category
        cursor.execute("""
            SELECT
                COALESCE(i.group_name, 'Uncategorized') as category,
                COUNT(DISTINCT i.item_name) as item_count,
                COALESCE(AVG(poi.unit_price), 0) as avg_cost,
                COALESCE(AVG(p.price), 0) as avg_selling_price,
                CASE
                    WHEN COALESCE(AVG(p.price), 0) > 0 THEN
                        ((COALESCE(AVG(p.price), 0) - COALESCE(AVG(poi.unit_price), 0)) / COALESCE(AVG(p.price), 0) * 100)
                    ELSE 0
                END as profit_margin_percent
            FROM items i
            LEFT JOIN purchase_order_items poi ON i.item_name = poi.item_name
            LEFT JOIN prices p ON i.item_name = p.item_name
            GROUP BY i.group_name
            ORDER BY profit_margin_percent DESC
        """)

        categories = []
        for row in cursor.fetchall():
            categories.append({
                "category": row['category'],
                "item_count": row['item_count'],
                "avg_cost": round(row['avg_cost'], 2),
                "avg_selling_price": round(row['avg_selling_price'], 2),
                "profit_margin_percent": round(row['profit_margin_percent'], 2)
            })

        conn.close()
        return {
            "by_item": items,
            "by_category": categories
        }

    except Exception as e:
        logging.error(f"Error calculating profit margins: {e}")
        raise HTTPException(status_code=500, detail=f"Error calculating profit margins: {str(e)}")

# ============================================================================
# INVENTORY FORECASTING
# ============================================================================

@app.get("/forecasting/demand-prediction")
async def predict_demand(
    item_name: Optional[str] = None,
    days_ahead: int = 30,
    current_user: User = Depends(get_current_user)
):
    """Predict future demand based on historical stock movements"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # Build item filter
        item_filter = ""
        params = []
        if item_name:
            item_filter = "WHERE sa.item_name = ?"
            params.append(item_name)

        # Get historical stock movements (last 90 days)
        cursor.execute(
            f"""SELECT
                   sa.item_name,
                   i.group_name,
                   DATE(sa.adjustment_date) as date,
                   SUM(CASE WHEN sa.adjustment_type = 'decrease' THEN sa.quantity ELSE 0 END) as quantity_out,
                   SUM(CASE WHEN sa.adjustment_type = 'increase' THEN sa.quantity ELSE 0 END) as quantity_in
               FROM stock_adjustments sa
               JOIN items i ON sa.item_name = i.item_name
               {item_filter}
               {'AND' if item_filter else 'WHERE'} sa.adjustment_date >= datetime('now', '-90 days')
               GROUP BY sa.item_name, i.group_name, DATE(sa.adjustment_date)
               ORDER BY sa.item_name, date""",
            params
        )

        movements = cursor.fetchall()

        # Calculate average daily consumption per item
        item_stats = {}
        for row in movements:
            item = row['item_name']
            if item not in item_stats:
                item_stats[item] = {
                    'item_name': item,
                    'group_name': row['group_name'],
                    'total_out': 0,
                    'total_in': 0,
                    'days_tracked': 0
                }
            item_stats[item]['total_out'] += row['quantity_out']
            item_stats[item]['total_in'] += row['quantity_in']
            item_stats[item]['days_tracked'] += 1

        # Calculate predictions
        predictions = []
        for item, stats in item_stats.items():
            avg_daily_consumption = stats['total_out'] / max(stats['days_tracked'], 1)
            predicted_demand = avg_daily_consumption * days_ahead

            # Get current stock
            cursor.execute("SELECT quantity, reorder_level FROM items WHERE item_name = ?", (item,))
            current = cursor.fetchone()
            current_stock = current['quantity'] if current else 0
            reorder_level = current['reorder_level'] if current and current['reorder_level'] else 0

            # Calculate stock depletion date
            days_until_depletion = (current_stock / avg_daily_consumption) if avg_daily_consumption > 0 else 999

            # Determine urgency
            if days_until_depletion < 7:
                urgency = 'critical'
            elif days_until_depletion < 14:
                urgency = 'high'
            elif days_until_depletion < 30:
                urgency = 'medium'
            else:
                urgency = 'low'

            predictions.append({
                'item_name': item,
                'group_name': stats['group_name'],
                'current_stock': current_stock,
                'avg_daily_consumption': round(avg_daily_consumption, 2),
                'predicted_demand_next_30_days': round(predicted_demand, 2),
                'days_until_depletion': round(days_until_depletion, 1) if days_until_depletion < 999 else None,
                'reorder_level': reorder_level,
                'should_reorder': current_stock <= reorder_level or days_until_depletion < 14,
                'urgency': urgency,
                'tracking_period_days': stats['days_tracked']
            })

        # Sort by urgency
        urgency_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3}
        predictions.sort(key=lambda x: urgency_order[x['urgency']])

        conn.close()
        return predictions

    except Exception as e:
        logging.error(f"Error predicting demand: {e}")
        raise HTTPException(status_code=500, detail=f"Error predicting demand: {str(e)}")


@app.get("/forecasting/reorder-recommendations")
async def get_reorder_recommendations(
    current_user: User = Depends(get_current_user)
):
    """Get items that should be reordered based on forecasting"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # Get all items with their current stock and reorder levels
        cursor.execute("""
            SELECT
                i.item_name,
                i.quantity as current_stock,
                i.reorder_level,
                i.group_name,
                COALESCE(AVG(sp.unit_price), 0) as avg_unit_price,
                COALESCE(AVG(sp.lead_time_days), 7) as avg_lead_time
            FROM items i
            LEFT JOIN supplier_products sp ON i.item_name = sp.item_name AND sp.is_available = 1
            GROUP BY i.item_name, i.quantity, i.reorder_level, i.group_name
        """)

        items = cursor.fetchall()

        recommendations = []
        for item in items:
            # Calculate historical consumption
            cursor.execute("""
                SELECT
                    COALESCE(AVG(CASE WHEN adjustment_type = 'decrease' THEN quantity ELSE 0 END), 0) as avg_consumption
                FROM stock_adjustments
                WHERE item_name = ?
                    AND adjustment_date >= datetime('now', '-30 days')
            """, (item['item_name'],))

            consumption_data = cursor.fetchone()
            avg_daily_consumption = consumption_data['avg_consumption'] if consumption_data else 0

            # Calculate recommended order quantity
            # Safety stock = avg daily consumption * lead time * 1.5 (safety factor)
            lead_time = item['avg_lead_time']
            safety_stock = avg_daily_consumption * lead_time * 1.5

            # Reorder quantity = (avg daily consumption * lead time) + safety stock - current stock
            reorder_quantity = max(0, (avg_daily_consumption * lead_time) + safety_stock - item['current_stock'])

            # Only recommend if below reorder level or will run out soon
            days_until_stockout = (item['current_stock'] / avg_daily_consumption) if avg_daily_consumption > 0 else 999

            if item['current_stock'] <= item['reorder_level'] or days_until_stockout < lead_time:
                recommendations.append({
                    'item_name': item['item_name'],
                    'group_name': item['group_name'],
                    'current_stock': item['current_stock'],
                    'reorder_level': item['reorder_level'],
                    'recommended_order_qty': round(reorder_quantity),
                    'avg_daily_consumption': round(avg_daily_consumption, 2),
                    'avg_lead_time_days': round(lead_time),
                    'days_until_stockout': round(days_until_stockout, 1) if days_until_stockout < 999 else None,
                    'estimated_cost': round(reorder_quantity * item['avg_unit_price'], 2),
                    'priority': 'critical' if days_until_stockout < 7 else 'high' if days_until_stockout < 14 else 'medium'
                })

        # Sort by priority
        priority_order = {'critical': 0, 'high': 1, 'medium': 2}
        recommendations.sort(key=lambda x: priority_order[x['priority']])

        conn.close()
        return recommendations

    except Exception as e:
        logging.error(f"Error getting reorder recommendations: {e}")
        raise HTTPException(status_code=500, detail=f"Error getting reorder recommendations: {str(e)}")


@app.get("/forecasting/stock-trends")
async def get_stock_trends(
    item_name: Optional[str] = None,
    days: int = 30,
    current_user: User = Depends(get_current_user)
):
    """Get stock level trends over time"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # Build item filter
        item_filter = ""
        params = [days]
        if item_name:
            item_filter = "AND h.item_name = ?"
            params.append(item_name)

        # Get historical stock levels from history
        cursor.execute(
            f"""SELECT
                   h.item_name,
                   DATE(h.timestamp) as date,
                   h.quantity,
                   h.action
               FROM history h
               WHERE h.timestamp >= datetime('now', '-? days')
               {item_filter}
               ORDER BY h.item_name, h.timestamp""",
            params
        )

        history = cursor.fetchall()

        # Group by item and date
        trends = {}
        for row in history:
            item = row['item_name']
            date = row['date']

            if item not in trends:
                trends[item] = {}

            if date not in trends[item]:
                trends[item][date] = {
                    'date': date,
                    'quantity': row['quantity'],
                    'actions': []
                }

            trends[item][date]['actions'].append(row['action'])

        # Format response
        result = []
        for item, dates in trends.items():
            result.append({
                'item_name': item,
                'trend_data': list(dates.values())
            })

        conn.close()
        return result

    except Exception as e:
        logging.error(f"Error getting stock trends: {e}")
        raise HTTPException(status_code=500, detail=f"Error getting stock trends: {str(e)}")


@app.get("/forecasting/seasonal-analysis")
async def get_seasonal_analysis(
    current_user: User = Depends(get_current_user)
):
    """Analyze seasonal patterns in inventory movement"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # Get movements by month for the past year
        cursor.execute("""
            SELECT
                sa.item_name,
                strftime('%m', sa.adjustment_date) as month,
                strftime('%Y', sa.adjustment_date) as year,
                SUM(CASE WHEN sa.adjustment_type = 'decrease' THEN sa.quantity ELSE 0 END) as total_out,
                COUNT(*) as movement_count
            FROM stock_adjustments sa
            WHERE sa.adjustment_date >= datetime('now', '-365 days')
            GROUP BY sa.item_name, month, year
            ORDER BY sa.item_name, year, month
        """)

        movements = cursor.fetchall()

        # Analyze patterns
        item_patterns = {}
        for row in movements:
            item = row['item_name']
            month = int(row['month'])

            if item not in item_patterns:
                item_patterns[item] = {
                    'item_name': item,
                    'monthly_data': {},
                    'peak_month': None,
                    'low_month': None,
                    'avg_monthly_movement': 0
                }

            if month not in item_patterns[item]['monthly_data']:
                item_patterns[item]['monthly_data'][month] = {
                    'month': month,
                    'total_movement': 0,
                    'count': 0
                }

            item_patterns[item]['monthly_data'][month]['total_movement'] += row['total_out']
            item_patterns[item]['monthly_data'][month]['count'] += 1

        # Calculate peaks and averages
        result = []
        month_names = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

        for item, data in item_patterns.items():
            if data['monthly_data']:
                monthly_totals = {m: d['total_movement'] for m, d in data['monthly_data'].items()}
                peak_month = max(monthly_totals, key=monthly_totals.get)
                low_month = min(monthly_totals, key=monthly_totals.get)
                avg_movement = sum(monthly_totals.values()) / len(monthly_totals)

                # Calculate seasonality index (peak / average)
                seasonality_index = (monthly_totals[peak_month] / avg_movement) if avg_movement > 0 else 1

                result.append({
                    'item_name': item,
                    'peak_month': month_names[peak_month],
                    'peak_month_movement': monthly_totals[peak_month],
                    'low_month': month_names[low_month],
                    'low_month_movement': monthly_totals[low_month],
                    'avg_monthly_movement': round(avg_movement, 2),
                    'seasonality_index': round(seasonality_index, 2),
                    'is_seasonal': seasonality_index > 1.5,
                    'monthly_breakdown': [
                        {
                            'month': month_names[m],
                            'movement': d['total_movement']
                        }
                        for m, d in sorted(data['monthly_data'].items())
                    ]
                })

        # Sort by seasonality index
        result.sort(key=lambda x: x['seasonality_index'], reverse=True)

        conn.close()
        return result

    except Exception as e:
        logging.error(f"Error analyzing seasonal patterns: {e}")
        raise HTTPException(status_code=500, detail=f"Error analyzing seasonal patterns: {str(e)}")

# ============================================================================
# SUPPLIER PERFORMANCE METRICS
# ============================================================================

@app.get("/suppliers/{supplier_id}/performance")
async def get_supplier_performance(
    supplier_id: int,
    current_user: User = Depends(get_current_user)
):
    """Get comprehensive performance metrics for a supplier"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # Verify supplier exists
        cursor.execute("SELECT name FROM suppliers WHERE id = ?", (supplier_id,))
        supplier = cursor.fetchone()
        if not supplier:
            raise HTTPException(status_code=404, detail="Supplier not found")

        # Get all purchase orders from this supplier
        cursor.execute("""
            SELECT
                id,
                order_date,
                expected_delivery_date,
                actual_delivery_date,
                status,
                total_amount
            FROM purchase_orders
            WHERE supplier_id = ?
        """, (supplier_id,))

        orders = cursor.fetchall()

        # Calculate on-time delivery rate
        total_completed = 0
        on_time_deliveries = 0
        total_late_days = 0

        for order in orders:
            if order['status'] == 'received' and order['expected_delivery_date'] and order['actual_delivery_date']:
                total_completed += 1
                expected = datetime.fromisoformat(order['expected_delivery_date'])
                actual = datetime.fromisoformat(order['actual_delivery_date'])
                diff = (actual - expected).days

                if diff <= 0:
                    on_time_deliveries += 1
                else:
                    total_late_days += diff

        on_time_rate = (on_time_deliveries / total_completed * 100) if total_completed > 0 else 0
        avg_delay_days = (total_late_days / (total_completed - on_time_deliveries)) if (total_completed - on_time_deliveries) > 0 else 0

        # Calculate price competitiveness
        cursor.execute("""
            SELECT
                sp.item_name,
                sp.unit_price as supplier_price,
                (SELECT AVG(unit_price) FROM supplier_products WHERE item_name = sp.item_name AND is_available = 1) as market_avg
            FROM supplier_products sp
            WHERE sp.supplier_id = ? AND sp.is_available = 1
        """, (supplier_id,))

        price_data = cursor.fetchall()
        competitive_count = 0
        total_items = len(price_data)

        for item in price_data:
            if item['market_avg'] and item['supplier_price'] <= item['market_avg']:
                competitive_count += 1

        price_competitiveness = (competitive_count / total_items * 100) if total_items > 0 else 0

        # Calculate order fulfillment statistics
        total_orders = len(orders)
        pending_orders = sum(1 for o in orders if o['status'] == 'pending')
        confirmed_orders = sum(1 for o in orders if o['status'] == 'confirmed')
        shipped_orders = sum(1 for o in orders if o['status'] == 'shipped')
        received_orders = sum(1 for o in orders if o['status'] == 'received')
        cancelled_orders = sum(1 for o in orders if o['status'] == 'cancelled')

        fulfillment_rate = (received_orders / (total_orders - cancelled_orders) * 100) if (total_orders - cancelled_orders) > 0 else 0
        cancellation_rate = (cancelled_orders / total_orders * 100) if total_orders > 0 else 0

        # Calculate total spending and average order value
        total_spent = sum(o['total_amount'] for o in orders if o['status'] == 'received')
        avg_order_value = (total_spent / received_orders) if received_orders > 0 else 0

        # Get lead time statistics
        cursor.execute("""
            SELECT AVG(lead_time_days) as avg_lead_time
            FROM supplier_products
            WHERE supplier_id = ? AND is_available = 1
        """, (supplier_id,))

        lead_time_data = cursor.fetchone()
        avg_lead_time = lead_time_data['avg_lead_time'] if lead_time_data and lead_time_data['avg_lead_time'] else 0

        # Get supplier rating
        cursor.execute("SELECT rating FROM suppliers WHERE id = ?", (supplier_id,))
        rating_data = cursor.fetchone()
        supplier_rating = rating_data['rating'] if rating_data and rating_data['rating'] else 0

        # Calculate overall performance score (0-100)
        # Weighted: On-time 30%, Price 25%, Fulfillment 25%, Rating 20%
        performance_score = (
            (on_time_rate * 0.30) +
            (price_competitiveness * 0.25) +
            (fulfillment_rate * 0.25) +
            (supplier_rating * 20 * 0.20)  # Convert rating from 1-5 to percentage
        )

        conn.close()

        return {
            'supplier_id': supplier_id,
            'supplier_name': supplier['name'],
            'overall_performance_score': round(performance_score, 2),
            'delivery_performance': {
                'on_time_delivery_rate': round(on_time_rate, 2),
                'total_completed_orders': total_completed,
                'on_time_deliveries': on_time_deliveries,
                'late_deliveries': total_completed - on_time_deliveries,
                'avg_delay_days': round(avg_delay_days, 2)
            },
            'price_performance': {
                'price_competitiveness_score': round(price_competitiveness, 2),
                'competitive_items': competitive_count,
                'total_items_supplied': total_items
            },
            'order_statistics': {
                'total_orders': total_orders,
                'pending_orders': pending_orders,
                'confirmed_orders': confirmed_orders,
                'shipped_orders': shipped_orders,
                'received_orders': received_orders,
                'cancelled_orders': cancelled_orders,
                'fulfillment_rate': round(fulfillment_rate, 2),
                'cancellation_rate': round(cancellation_rate, 2)
            },
            'financial_metrics': {
                'total_spent': round(total_spent, 2),
                'avg_order_value': round(avg_order_value, 2),
                'avg_lead_time_days': round(avg_lead_time, 2)
            },
            'quality_rating': supplier_rating
        }

    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Error getting supplier performance: {e}")
        raise HTTPException(status_code=500, detail=f"Error getting supplier performance: {str(e)}")


@app.get("/suppliers/performance/comparison")
async def compare_supplier_performance(
    current_user: User = Depends(get_current_user)
):
    """Compare performance across all active suppliers"""
    try:
        conn = db.get_connection()
        cursor = conn.cursor()

        # Get all active suppliers
        cursor.execute("SELECT id, name FROM suppliers WHERE is_active = 1")
        suppliers = cursor.fetchall()

        comparisons = []

        for supplier in suppliers:
            # Get basic performance metrics
            cursor.execute("""
                SELECT
                    COUNT(*) as total_orders,
                    SUM(CASE WHEN status = 'received' THEN 1 ELSE 0 END) as completed_orders,
                    SUM(CASE WHEN status = 'received' THEN total_amount ELSE 0 END) as total_spent
                FROM purchase_orders
                WHERE supplier_id = ?
            """, (supplier['id'],))

            stats = cursor.fetchone()

            # Calculate on-time rate
            cursor.execute("""
                SELECT
                    COUNT(*) as total_on_time
                FROM purchase_orders
                WHERE supplier_id = ?
                    AND status = 'received'
                    AND actual_delivery_date IS NOT NULL
                    AND expected_delivery_date IS NOT NULL
                    AND actual_delivery_date <= expected_delivery_date
            """, (supplier['id'],))

            on_time_data = cursor.fetchone()
            on_time_rate = (on_time_data['total_on_time'] / stats['completed_orders'] * 100) if stats['completed_orders'] > 0 else 0

            # Get average price rank
            cursor.execute("""
                SELECT COUNT(*) as items_supplied
                FROM supplier_products
                WHERE supplier_id = ? AND is_available = 1
            """, (supplier['id'],))

            items_data = cursor.fetchone()

            # Get supplier rating
            cursor.execute("SELECT rating FROM suppliers WHERE id = ?", (supplier['id'],))
            rating_data = cursor.fetchone()
            rating = rating_data['rating'] if rating_data and rating_data['rating'] else 0

            comparisons.append({
                'supplier_id': supplier['id'],
                'supplier_name': supplier['name'],
                'total_orders': stats['total_orders'],
                'completed_orders': stats['completed_orders'],
                'total_spent': round(stats['total_spent'], 2) if stats['total_spent'] else 0,
                'on_time_delivery_rate': round(on_time_rate, 2),
                'items_supplied': items_data['items_supplied'],
                'quality_rating': rating,
                'avg_order_value': round(stats['total_spent'] / stats['completed_orders'], 2) if stats['completed_orders'] > 0 else 0
            })

        # Sort by total spent (highest first)
        comparisons.sort(key=lambda x: x['total_spent'], reverse=True)

        conn.close()
        return comparisons

    except Exception as e:
        logging.error(f"Error comparing supplier performance: {e}")
        raise HTTPException(status_code=500, detail=f"Error comparing supplier performance: {str(e)}")

# ============================================================================
# Health Check
# ============================================================================

@app.get("/health")
async def health_check():
    """API health check"""
    return {"status": "healthy", "message": "Inventory Management API is running"}

@app.get("/")
async def root():
    """Root endpoint"""
    return {
        "message": "Inventory Management System API",
        "version": "1.0.0",
        "docs": "/docs",
        "health": "/health"
    }

# ============================================================================
# Run Server
# ============================================================================

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8001))
    logging.info(f"Starting server on port {port}")
    uvicorn.run(app, host="0.0.0.0", port=port, log_level="info")
